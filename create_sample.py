from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = 'Products'

# Add headers
headers = ['Product Name', 'Quantity', 'Purchasing Price', 'Sale Price', 'Category', 'Expiry Date', 'ISBN', 'Production Date', 'Author', 'Supplier']
for col_num, header in enumerate(headers, 1):
    ws.cell(row=1, column=col_num, value=header)

# Add sample data
sample_data = [
    ['Sample Book 1', 10, 15.50, 25.00, 'Fiction', '2025-12-31', '9781234567890', '2024-01-15', 'John Doe', 'ABC Publishers'],
    ['Sample Book 2', 5, 12.75, 20.00, 'Non-Fiction', '', '9780987654321', '2024-02-20', 'Jane Smith', 'XYZ Books'],
    ['Sample Magazine', 20, 5.00, 8.50, 'Magazine', '2024-12-31', '9781122334455', '2024-03-01', '', 'News Corp']
]

for row_num, row_data in enumerate(sample_data, 2):
    for col_num, value in enumerate(row_data, 1):
        ws.cell(row=row_num, column=col_num, value=value)

wb.save('my_app/static/sample_products_import.xlsx')
print('Sample Excel file created successfully!')