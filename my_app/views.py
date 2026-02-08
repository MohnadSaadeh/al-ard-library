from django.utils.translation import gettext as _

def stock_products_report(request):
    products = models.Product.objects.filter(quantity__gt=0)
    context = {
        'products': products,
    }
    return render(request, 'stock_products_report.html', context)

def empty_products_report(request):
    products = models.Product.objects.filter(quantity=0)
    context = {
        'products': products,
    }
    return render(request, 'empty_products_report.html', context)
# AJAX endpoint: add product to sale cart by ISBN
from itertools import product
from django.views.decorators.csrf import csrf_exempt
import json

@csrf_exempt
def add_product_to_sale_cart_by_isbn(request):
    """
    AJAX endpoint: receive 'isbn' (POST, JSON) and add product to sale cart by ISBN.
    Returns JSON with status, grand_total, and updated cart items.
    """
    try:
        if request.method != 'POST':
            return JsonResponse({'status': 'error', 'message': _('Invalid method')}, status=405)

        # Accept JSON or form POST
        if request.content_type == 'application/json':
            try:
                data = json.loads(request.body.decode('utf-8'))
                isbn = data.get('isbn', '').strip()
            except Exception:
                isbn = ''
        else:
            isbn = request.POST.get('isbn', '').strip()

        if not isbn:
            return JsonResponse({'status': 'error', 'message': _('No ISBN provided')}, status=400)

        try:
            product = models.Product.objects.get(isbn=isbn)
        except models.Product.DoesNotExist:
            return JsonResponse({'status': 'not_found', 'message': _('Product not found')})

        # Use add_product_to_sale logic for validation and addition
        if product.quantity <= 0:
            return JsonResponse({'status': 'error', 'message': _('Cannot sell "%(product)s": out of stock.') % {'product': product.product_name}})

        cart = _get_sale_cart(request)
        for item in cart:
            if int(item.get('product_id')) == int(product.id):
                new_quantity = int(item.get('quantity', 0)) + 1
                if new_quantity > product.quantity:
                    return JsonResponse({'status': 'error', 'message': _('Cannot sell more than available stock for "%(product)s".') % {'product': product.product_name}})
                item['quantity'] = new_quantity
                try:
                    item['total_price'] = int(item['quantity']) * float(item.get('sale_price', 0))
                except Exception:
                    item['total_price'] = 0
                break
        else:
            if product.quantity < 1:
                return JsonResponse({'status': 'error', 'message': _('Cannot sell more than available stock for "%(product)s".') % {'product': product.product_name}})
            sale_price = float(product.sale_price) if product.sale_price is not None else 0.0
            cart.append({
                'product_name': product.product_name,
                'product_id': product.id,
                'quantity': 1,
                'sale_price': sale_price,
                'total_price': sale_price,
            })

        _save_sale_cart(request, cart)
        grand_total = sum(float(item.get('total_price', 0)) for item in cart)
        return JsonResponse({'status': 'ok', 'grand_total': grand_total, 'items': cart})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': f'Internal error: {str(e)}'})
from django.shortcuts import render ,redirect
from . import models
from .models import Purchase, Sale_order
from django.contrib import messages
import bcrypt
from datetime import datetime , timedelta
import datetime
from django.db.models import Sum
from django.db.models import Q
from django.db import transaction
from django.db.models.functions import ExtractWeekDay
from django.http import JsonResponse
from . import validations
from django.core.paginator import Paginator
from django.utils.translation import gettext as _
import secrets
from django.utils import timezone
from django.core.mail import send_mail
from django.conf import settings



# Session-backed cart helpers (per-user carts stored in session)
def _get_sale_cart(request):
    return request.session.get('sale_order', [])


def _save_sale_cart(request, cart):
    request.session['sale_order'] = cart
    request.session.modified = True


def _get_purchase_cart(request):
    return request.session.get('purchases_order', [])


def _save_purchase_cart(request, cart):
    request.session['purchases_order'] = cart
    request.session.modified = True


def delete_product_from_purchase(request):
    """
    AJAX endpoint: delete a product from purchase cart by product_id
    Returns JSON with updated grand_total and cart items
    """
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid method'}, status=405)
    
    product_id = request.POST.get('product_id')
    if not product_id:
        return JsonResponse({'status': 'error', 'message': 'No product_id provided'}, status=400)
    
    try:
        product_id = int(product_id)
    except (ValueError, TypeError):
        return JsonResponse({'status': 'error', 'message': 'Invalid product_id'}, status=400)
    
    cart = _get_purchase_cart(request)
    # Find and remove the item by product_id
    original_length = len(cart)
    cart = [item for item in cart if int(item.get('product_id')) != product_id]
    
    if len(cart) == original_length:
        # Item not found
        return JsonResponse({'status': 'error', 'message': 'Product not found in cart'}, status=404)
    
    _save_purchase_cart(request, cart)
    
    # Recalculate grand total
    grand_total = sum(float(item.get('total_price', 0)) for item in cart)
    
    return JsonResponse({'status': 'ok', 'grand_total': grand_total, 'items': cart})

def update_purchase_cart(request):
    """
    Persist edited purchase cart rows to session.
    Accepts JSON payload: { items: [ {product_id, quantity, purchase_price} ] }
    Returns updated items and grand_total.
    """
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid method'}, status=405)

    items_payload = None
    if request.content_type == 'application/json':
        import json
        try:
            body = json.loads(request.body or '{}')
            items_payload = body.get('items')
        except Exception:
            return JsonResponse({'status': 'error', 'message': _('Invalid payload')}, status=400)
    else:
        items_str = request.POST.get('items')
        if items_str:
            import json
            try:
                items_payload = json.loads(items_str)
            except Exception:
                items_payload = None

    if not isinstance(items_payload, list):
        return JsonResponse({'status': 'error', 'message': _('No items provided')}, status=400)

    new_cart = []
    for it in items_payload:
        product_id = it.get('product_id')
        quantity = it.get('quantity')
        purchase_price = it.get('purchase_price')
        try:
            product_id = int(product_id)
        except (ValueError, TypeError):
            return JsonResponse({'status': 'error', 'message': _('Invalid product_id')}, status=400)
        try:
            quantity = int(quantity)
        except (ValueError, TypeError):
            quantity = 0
        try:
            purchase_price = float(purchase_price)
        except (ValueError, TypeError):
            purchase_price = 0.0

        if quantity <= 0:
            return JsonResponse({'status': 'error', 'message': _('Quantity must be positive')}, status=400)

        try:
            prod = models.Product.objects.get(id=product_id)
        except models.Product.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': _('Product not found')}, status=404)

        total_price = quantity * purchase_price
        new_cart.append({
            'product_name': prod.product_name,
            'product_id': product_id,
            'quantity': quantity,
            'purchase_price': purchase_price,
            'total_price': total_price,
        })

    _save_purchase_cart(request, new_cart)

    grand_total = sum(float(item.get('total_price', 0)) for item in new_cart)
    return JsonResponse({'status': 'ok', 'grand_total': grand_total, 'items': new_cart})


# to display the sign-in page
def about_us(request):
    return render(request, 'about_us.html')




def display_homepage(request):
    return render(request, 'sign-in.html')

def index(request):
    if 'employee_id' in request.session:
        context = {   
            'segment': index,        
            'sixmonthesproducts': models.get_six_monthes_products(), # MAI ******
            'near_expiry':models.get_six_monthes(),   # MAI ******        
            'out_stock':models.out_of_stock(),
            'count':models.count_out_stock(),
            'today_sale_order':models.today_sale_orders(), #MAI *******
            'today_purchases_order' : models.today_purchases(),
            'employee': models.get_employee_by_id(request.session['employee_id']),

            # 'chart_data':models.chart_sales_orders()
            # 'saturday' : models.Sale_order.objects.filter(created_at__contains__week_day=7).count(),
            'monday' : models.Sale_order.objects.filter(created_at__week_day=2).count(),
            'tuseday' : models.Sale_order.objects.filter(created_at__week_day=3).count(),
            'wednesday' : models.Sale_order.objects.filter(created_at__week_day=4).count(),
            'thurseday' : models.Sale_order.objects.filter(created_at__week_day=5).count(),
            'friday' : models.Sale_order.objects.filter(created_at__week_day=6).count(),
            'saturday' : models.Sale_order.objects.filter(created_at__week_day=7).count(),
            'sunday' : models.Sale_order.objects.filter(created_at__week_day=1).count(),

            # 'employees' : active_employees,
            

            }
        return render(request , 'index.html' , context)
    else:
        return redirect('/')
    

def sign_up(request):
    return render(request , 'sign-up.html' )

# if anyone want to Sign is
def sign_in(request):
    # Translate messages for user feedback
    if request.method == 'POST':
        if request.POST['account_type'] == "1":  # Employee
            errors = models.Employee.objects.login_employee_validator(request.POST)
            if len(errors) > 0:
                for key, value in errors.items():
                    messages.error(request, _(value))  # Already translated
                return redirect('/')
            else:
                employee_email = request.POST['email']
                employee_password = request.POST['password']
                employee = models.Employee.objects.filter(email=employee_email)
                if employee:
                    employee_user = employee[0]
                    if bcrypt.checkpw(employee_password.encode(), employee_user.password.encode()):
                        request.session['employee_id'] = employee_user.id
                        #---------------is Active -----------------------
                        employee = models.Employee.objects.get(id=request.session['employee_id'])
                        employee.is_active = True
                        employee.save()
                        #---------------is Active -----------------------
                        return redirect('/index')
                    else:
                        messages.error(request, _("Incorrect Password"))  # Already translated
                        # messages.error(request, value , extra_tags = 'admin_login' )
                        return redirect('/')
                messages.error(request, _("Email is incorrect"))  # Already translated
                return redirect('/')
        # if he is a MANAGER
        else:
            if request.POST['account_type'] == "2" : # Manager
                errors = models.Manager.objects.login_manager_validator(request.POST)
                if len(errors) > 0:
                    for key, value in errors.items():
                        messages.error(request, value )
                    return redirect('/')
                else:
                    manager_email = request.POST['email'] # here we get the email thet ENSERTED
                    manager_password = request.POST['password'] # here we get the password thet ENSERTED
                    manager = models.Manager.objects.filter(email=manager_email) # here we get the MANAGER by the email from DB
                    if manager: # here we check if the MANAGER exist
                        manager_user = manager[0] # here we get the MANAGER from the list
                        #ADDING MANAGER FROME ADMIN
                        if manager_password == manager_user.password :
                            request.session['manager_id'] = manager_user.id
                            return redirect('/index')
                        else:
                            messages.error(request, _("Incorrect Password"))  # Already translated
                            # messages.error(request, value , extra_tags = 'admin_login' )
                            return redirect('/')
                    messages.error(request, _("Email is incorrect"))  # Already translated
    else:
        return redirect('/')

                #     if bcrypt.checkpw(manager_password.encode(), manager_user.password.encode()): # here we chick the password 
                #         request.session['manager_id'] = manager_user.id
                #         return redirect('/index')
                #     else:
                #         messages.error(request, "Incorrect Password")
                #         # messages.error(request, value , extra_tags = 'admin_login' )
                #         return redirect('/')
                # messages.error(request, "Email is incorrect")
                # return redirect('/')
    return redirect('/')

def change_password(request):
    if not ('employee_id' in request.session or 'manager_id' in request.session):
        return redirect('/')
    
    if request.method == 'POST':
        old_password = request.POST.get('old_password')
        new_password = request.POST.get('new_password')
        confirm_password = request.POST.get('confirm_password')
        
        if not old_password or not new_password or not confirm_password:
            messages.error(request, _("All fields are required"))
            return redirect('/change_password')
        
        if new_password != confirm_password:
            messages.error(request, _("New passwords do not match"))
            return redirect('/change_password')
        
        # Basic validation
        if len(new_password) < 8:
            messages.error(request, _("Password must be at least 8 characters"))
            return redirect('/change_password')
        
        if 'employee_id' in request.session:
            employee = models.Employee.objects.get(id=request.session['employee_id'])
            if not bcrypt.checkpw(old_password.encode(), employee.password.encode()):
                messages.error(request, _("Incorrect old password"))
                return redirect('/change_password')
            pw_hash = bcrypt.hashpw(new_password.encode(), bcrypt.gensalt()).decode()
            employee.password = pw_hash
            employee.confirm_password = pw_hash
            employee.save()
        elif 'manager_id' in request.session:
            manager = models.Manager.objects.get(id=request.session['manager_id'])
            # For managers, check if password is hashed or plain
            try:
                if bcrypt.checkpw(old_password.encode(), manager.password.encode()):
                    is_hashed = True
                else:
                    is_hashed = False
            except:
                is_hashed = False
            if not is_hashed and manager.password != old_password:
                messages.error(request, _("Incorrect old password"))
                return redirect('/change_password')
            pw_hash = bcrypt.hashpw(new_password.encode(), bcrypt.gensalt()).decode()
            manager.password = pw_hash
            manager.confirm_password = pw_hash
            manager.save()
        
        messages.success(request, _("Password changed successfully"))
        return redirect('/index')
    else:
        return render(request, 'change_password.html')

def forget_password(request):
    if request.method == 'POST':
        email = request.POST.get('email', '').strip()
        if not email:
            messages.error(request, _("Email is required"))
            return redirect('/forget_password')
        
        # Check if employee or manager
        employee = models.Employee.objects.filter(email=email).first()
        manager = models.Manager.objects.filter(email=email).first()
        
        if not employee and not manager:
            messages.error(request, _("Email not found"))
            return redirect('/forget_password')
        
        user_type = 'employee' if employee else 'manager'
        user_id = employee.id if employee else manager.id
        
        # Generate token
        token = secrets.token_urlsafe(32)
        expires_at = timezone.now() + timedelta(hours=1)
        
        models.PasswordResetToken.objects.create(
            user_type=user_type,
            user_id=user_id,
            token=token,
            expires_at=expires_at
        )
        
        # Send email
        reset_url = f"http://127.0.0.1:8000/reset_password/{token}"
        subject = _("Password Reset Request")
        message = _("Click the link to reset your password: ") + reset_url
        from_email = settings.EMAIL_HOST_USER
        
        try:
            send_mail(subject, message, from_email, [email])
            messages.success(request, _("Reset link sent to your email"))
        except Exception as e:
            messages.error(request, _("Failed to send email: %(error)s") % {'error': str(e)})
        
        return redirect('/')
    
    return render(request, 'forget_password.html')

def reset_password(request, token):
    try:
        reset_token = models.PasswordResetToken.objects.get(token=token, expires_at__gt=timezone.now())
    except models.PasswordResetToken.DoesNotExist:
        messages.error(request, _("Invalid or expired token"))
        return redirect('/')
    
    if request.method == 'POST':
        new_password = request.POST.get('new_password', '').strip()
        confirm_password = request.POST.get('confirm_password', '').strip()
        
        if not new_password or not confirm_password:
            messages.error(request, _("All fields are required"))
            return redirect(request.path)
        
        if new_password != confirm_password:
            messages.error(request, _("Passwords do not match"))
            return redirect(request.path)
        
        if len(new_password) < 8:
            messages.error(request, _("Password must be at least 8 characters"))
            return redirect(request.path)
        
        # Hash and save
        pw_hash = bcrypt.hashpw(new_password.encode(), bcrypt.gensalt()).decode()
        
        if reset_token.user_type == 'employee':
            employee = models.Employee.objects.get(id=reset_token.user_id)
            employee.password = pw_hash
            employee.confirm_password = pw_hash
            employee.save()
        else:
            manager = models.Manager.objects.get(id=reset_token.user_id)
            manager.password = pw_hash
            manager.confirm_password = pw_hash
            manager.save()
        
        reset_token.delete()
        messages.success(request, _("Password reset successfully"))
        return redirect('/')
    
    return render(request, 'reset_password.html')

def logout(request):
    request.session.clear()
    return redirect('/')
def emp_logout(request):
    employee = models.Employee.objects.get(id=request.session['employee_id'])
    employee.is_active = False
    employee.save()
    request.session.clear()
    return redirect('/')

def display_stock_for_manager(request):
    context={
        'products': models.get_all_products(),
        'today': datetime.date.today(),
        'expiry_range': datetime.date.today() + timedelta(days=6*30),
    }
    return render(request, 'stock_manager.html',context)



def add_new_employee(request):
    errors = models.Employee.objects.employee_validator(request.POST)
    if len(errors) > 0:
        for key, value in errors.items():
            messages.error(request, _(value))
        return redirect('/signup')
    else:
        # manager = request.session['manager_id']
        # manager = request.session['employee_id'] # or manager_id
        emp_f_name = request.POST['f_name']
        emp_l_name = request.POST['l_name']
        emp_emeil = request.POST['email']
        emp_DOB = request.POST['DOB']
        emp_password = request.POST['password']
        emp_conf_password = request.POST['c_password']
        #hash-----Passwords-------
        pw_hash = bcrypt.hashpw(emp_password.encode(), bcrypt.gensalt()).decode()
        pw_hash_confirm = bcrypt.hashpw(emp_conf_password.encode(), bcrypt.gensalt()).decode()
        #hash-----Passwords-------
        models.add_employee(emp_f_name, emp_l_name, emp_emeil, emp_DOB, pw_hash, pw_hash_confirm  )
    messages.success(request, _("Successfully added an employee!"), extra_tags = 'add_employee')
    return redirect('/signup')

def display_employees(request):
    # if the manad=egr not in the loged on
    # 'manager_id' or 'employee_id' in request.session:
    if 'employee_id' not in request.session:
        return redirect('/index')
    else:
        context = {
            'employees': models.get_all_employees(),
            'employee': models.get_employee_by_id(request.session['employee_id']),

        }
        return render (request, 'profile.html', context )






def add_new_product(request):
    errors = models.Product.objects.product_validator(request.POST)
    if len(errors) > 0:
        for key, value in errors.items():
            messages.error(request, _(value))
        return redirect('/employye_dashboard')
    else:
        employee = request.session['employee_id']
        product_name = request.POST['product_name']
        quantity = request.POST['quantity']
        purchasing_price = request.POST['purchasing_price']
        expiry_date = request.POST['expiry_date']
        supplier = request.POST['supplier']
        models.add_product(product_name, quantity, purchasing_price, expiry_date, supplier, employee)
    messages.success(request, _("Successfully added a product!"), extra_tags = 'add_product')
    return redirect('/employye_dashboard')



def display_sales(request):
    # calculate grand total for current sale order (session-backed)
    cart = _get_sale_cart(request)
    grand_total = sum(float(item.get('total_price', 0)) for item in cart)
    # paginate recent sales orders
    orders_qs = models.get_all_sales_orders()
    orders_page_number = request.GET.get('orders_page')
    orders_paginator = Paginator(orders_qs, 10)  # 10 orders per page
    orders_page = orders_paginator.get_page(orders_page_number)
    base_currency = models.CompanyProfile.objects.first().base_currency_id if models.CompanyProfile.objects.first() and models.CompanyProfile.objects.first().base_currency else None

    context = {
            'sale_order': cart,
            'products': models.get_all_products(),
            'orders' : orders_page,
            'employee': models.get_employee_by_id(request.session['employee_id']),
            'customers': models.Customer.objects.all(),
            'currencies': models.Currency.objects.all(),
            'grand_total': grand_total,
            'orders_paginator': orders_paginator,
            'base_currency': base_currency,
        }
    return render(request , 'sale_orders.html', context )

def display_purchases(request):
    # calculate grand total (session-backed)
    p_cart = _get_purchase_cart(request)
    grand_total = sum(float(item.get('total_price', 0)) for item in p_cart)
    # paginate recent purchase invoices
    invoices_qs = models.get_all_invoices()
    invoices_page_number = request.GET.get('invoices_page')
    invoices_paginator = Paginator(invoices_qs, 10)  # 10 invoices per page
    invoices_page = invoices_paginator.get_page(invoices_page_number)
    

    context = {
            'purchases_order': p_cart,
            'products': models.get_all_products(),#--------------------------------------------Mai
            'invoices' : invoices_page,
            'employee': models.get_employee_by_id(request.session['employee_id']),#--------------------------------------------Mai
            'grand_total': grand_total,
            'invoices_paginator': invoices_paginator,
            'suppliers': models.Supplier.objects.all(),
            'currencies': models.Currency.objects.all(),
        }
    return render(request , 'purchase_invoices.html' ,context)


def display_suppliers(request):
    # List all suppliers and provide a simple add form on the same page
    if 'employee_id' not in request.session:
        return redirect('/index')
    q = request.GET.get('q', '').strip()
    suppliers_qs = models.Supplier.objects.all()
    if q:
        suppliers_qs = suppliers_qs.filter(
            Q(name__icontains=q) |
            Q(phone__icontains=q) |
            Q(email__icontains=q) |
            Q(contact_info__icontains=q)
        )
    suppliers_qs = suppliers_qs.order_by('-created_at')
    # paginate results (12 per page)
    page_number = request.GET.get('page')
    paginator = Paginator(suppliers_qs, 12)
    suppliers_page = paginator.get_page(page_number)

    context = {
        'suppliers': suppliers_page,
        'employee': models.get_employee_by_id(request.session['employee_id']),
        'suppliers_paginator': paginator,
    }
    return render(request, 'suppliers.html', context)


def add_supplier(request):
    # Simple POST handler to create a supplier
    if request.method != 'POST':
        return redirect('/suppliers')
    name = request.POST.get('name', '').strip()
    phone = request.POST.get('phone', '').strip()
    email = request.POST.get('email', '').strip()
    contact_info = request.POST.get('contact_info', '').strip()

    if not name:
        messages.error(request, _('Supplier name is required.'))
        return redirect('/suppliers')

    try:
        models.Supplier.objects.create(name=name, phone=phone or None, email=email or None, contact_info=contact_info or None)
        messages.success(request, _('Successfully added supplier.'), extra_tags='add_supplier')
    except Exception as e:
        messages.error(request, _('Failed to add supplier: %(err)s') % {'err': str(e)})

    return redirect('/suppliers')


def edit_supplier(request, id):
    # Edit supplier (GET shows edit form, POST updates)
    try:
        supplier = models.Supplier.objects.get(id=id)
    except models.Supplier.DoesNotExist:
        messages.error(request, _('Supplier not found.'))
        return redirect('/suppliers')

    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        phone = request.POST.get('phone', '').strip()
        email = request.POST.get('email', '').strip()
        contact_info = request.POST.get('contact_info', '').strip()

        if not name:
            messages.error(request, _('Supplier name is required.'))
            return redirect(f'/suppliers/{id}/edit')

        supplier.name = name
        supplier.phone = phone or None
        supplier.email = email or None
        supplier.contact_info = contact_info or None
        supplier.save()
        messages.success(request, _('Supplier updated.'), extra_tags='edit_supplier')
        return redirect('/suppliers')

    # GET
    context = {
        'supplier': supplier,
        'employee': models.get_employee_by_id(request.session.get('employee_id')) if request.session.get('employee_id') else None,
    }
    return render(request, 'suppliers_edit.html', context)


def delete_supplier(request, id):
    # Only accept POST for delete
    if request.method != 'POST':
        return redirect('/suppliers')
    try:
        supplier = models.Supplier.objects.get(id=id)
        supplier.delete()
        messages.success(request, _('Supplier deleted.'), extra_tags='delete_supplier')
    except models.Supplier.DoesNotExist:
        messages.error(request, _('Supplier not found.'))
    except Exception as e:
        messages.error(request, _('Failed to delete supplier: %(err)s') % {'err': str(e)})
    return redirect('/suppliers')


def supplier_detail(request, id):
    # Show supplier info and list of purchases from this supplier
    if 'employee_id' not in request.session:
        return redirect('/index')
    try:
        supplier = models.Supplier.objects.get(id=id)
    except models.Supplier.DoesNotExist:
        messages.error(request, _('Supplier not found.'))
        return redirect('/suppliers')

    purchases_qs = models.Purchase.objects.filter(supplier=supplier).order_by('-created_at')
    # paginate purchases
    page_number = request.GET.get('page')
    paginator = Paginator(purchases_qs, 10)
    purchases_page = paginator.get_page(page_number)

    context = {
        'supplier': supplier,
        'purchases': purchases_page,
        'purchases_paginator': paginator,
        'employee': models.get_employee_by_id(request.session['employee_id']),
    }
    return render(request, 'supplier_detail.html', context)

def display_customers(request):
    # List all customers and provide add form on the same page (mirrors suppliers)
    if 'employee_id' not in request.session:
        return redirect('/index')
    q = request.GET.get('q', '').strip()
    customers_qs = models.Customer.objects.all()
    if q:
        customers_qs = customers_qs.filter(
            Q(name__icontains=q) |
            Q(phone__icontains=q) |
            Q(email__icontains=q) |
            Q(contact_info__icontains=q)
        )
    customers_qs = customers_qs.order_by('-created_at')
    page_number = request.GET.get('page')
    paginator = Paginator(customers_qs, 12)
    customers_page = paginator.get_page(page_number)

    context = {
        'customers': customers_page,
        'employee': models.get_employee_by_id(request.session['employee_id']),
        'customers_paginator': paginator,
    }
    return render(request, 'customers.html', context)


def add_customer(request):
    if request.method != 'POST':
        return redirect('/customers')
    name = request.POST.get('name', '').strip()
    phone = request.POST.get('phone', '').strip()
    email = request.POST.get('email', '').strip()
    contact_info = request.POST.get('contact_info', '').strip()

    if not name:
        messages.error(request, _('Customer name is required.'))
        return redirect('/customers')

    try:
        models.Customer.objects.create(name=name, phone=phone or None, email=email or None, contact_info=contact_info or None)
        messages.success(request, _('Successfully added customer.'), extra_tags='add_customer')
    except Exception as e:
        messages.error(request, _('Failed to add customer: %(err)s') % {'err': str(e)})

    return redirect('/customers')


def edit_customer(request, id):
    try:
        customer = models.Customer.objects.get(id=id)
    except models.Customer.DoesNotExist:
        messages.error(request, _('Customer not found.'))
        return redirect('/customers')

    if request.method == 'POST':
        customer.name = request.POST.get('name', '').strip() or customer.name
        customer.phone = request.POST.get('phone', '').strip() or customer.phone
        customer.email = request.POST.get('email', '').strip() or customer.email
        customer.contact_info = request.POST.get('contact_info', '').strip() or customer.contact_info
        customer.save()
        messages.success(request, _('Customer updated successfully.'))
        return redirect('/customers')

    context = {
        'customer': customer,
        'employee': models.get_employee_by_id(request.session['employee_id']),
    }
    return render(request, 'customers_edit.html', context)


def delete_customer(request, id):
    try:
        c = models.Customer.objects.get(id=id)
        c.delete()
        messages.success(request, _('Customer deleted successfully.'))
    except Exception:
        messages.error(request, _('Failed to delete customer.'))
    return redirect('/customers')


def customer_detail(request, id):
    try:
        customer = models.Customer.objects.get(id=id)
    except models.Customer.DoesNotExist:
        messages.error(request, _('Customer not found.'))
        return redirect('/customers')

    sales_qs = models.Sale_order.objects.filter(customer=customer).order_by('-created_at') 
    stock_out_vouchers_qs = models.Stock_Out_Voucher.objects.filter(customer=customer).order_by('-created_at')
    page_number = request.GET.get('page')
    paginator = Paginator(sales_qs, 10)
    sales_page = paginator.get_page(page_number)

    context = {
        'customer': customer,
        'sales': sales_page,
        'employee': models.get_employee_by_id(request.session['employee_id']),
        'sales_paginator': paginator,
        'stock_outs': stock_out_vouchers_qs,
    }
    return render(request, 'customer_detail.html', context)

def delete_product(request):
    models.delete_clicked_product(request)
    return redirect('/employye_dashboard')


# session-backed sale_order is stored per-user in request.session via helpers above
#____________________________________SALE___________________________________
def add_product_to_sale(request):
    # Support both normal POST and AJAX/JSON requests. When called via AJAX return JSON
    data = None
    if request.content_type == 'application/json':
        try:
            data = json.loads(request.body)
        except Exception:
            return JsonResponse({'status': 'error', 'message': _('Invalid payload')}, status=400)

    # extract fields from either JSON payload or form POST
    product_name = (data.get('product_name') if data else request.POST.get('product_name'))
    quantity = int((data.get('quantity') if data else request.POST.get('quantity', 0)) or 0)

    if not product_name or quantity <= 0:
        msg = _('Product name and positive quantity are required.')
        if request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.content_type == 'application/json':
            return JsonResponse({'status': 'error', 'message': msg}, status=400)
        messages.error(request, msg)
        return redirect('/sales')

    try:
        prod = models.Product.objects.get(product_name=product_name)
    except models.Product.DoesNotExist:
        msg = _('Product not found')
        if request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.content_type == 'application/json':
            return JsonResponse({'status': 'not_found', 'message': msg})
        messages.error(request, msg)
        return redirect('/sales')

    product_id = prod.id

    if prod.quantity <= 0:
        msg = _('Cannot sell "%(product)s": out of stock.') % {'product': product_name}
        if request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.content_type == 'application/json':
            return JsonResponse({'status': 'error', 'message': msg})
        messages.error(request, msg)
        return redirect('/sales')

    # Check if adding this quantity would exceed available stock
    cart = _get_sale_cart(request)
    current_cart_quantity = 0
    for item in cart:
        if int(item.get('product_id')) == int(product_id):
            current_cart_quantity = int(item.get('quantity', 0))
            break

    total_quantity_after_add = current_cart_quantity + quantity
    if total_quantity_after_add > prod.quantity:
        msg = _('Cannot sell more than available stock for "%(product)s". Available: %(available)d, requested total: %(total)d') % {
            'product': product_name,
            'available': prod.quantity,
            'total': total_quantity_after_add
        }
        if request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.content_type == 'application/json':
            return JsonResponse({'status': 'error', 'message': msg})
        messages.error(request, msg)
        return redirect('/sales')

    sale_price = float(prod.sale_price) if prod.sale_price is not None else 0.0
    total_price = quantity * sale_price
    # if product already in cart, increase quantity
    for item in cart:
        if int(item.get('product_id')) == int(product_id):
            item['quantity'] = int(item.get('quantity', 0)) + quantity
            try:
                item['total_price'] = int(item['quantity']) * float(item.get('sale_price', 0))
            except Exception:
                item['total_price'] = 0
            break
    else:
        cart.append({
            'product_name': product_name,
            'product_id': product_id,
            'quantity': quantity,
            'sale_price': sale_price,
            'total_price': total_price,
        })
    _save_sale_cart(request, cart)

    # compute grand total
    grand_total = sum(float(item.get('total_price', 0)) for item in cart)
    if request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.content_type == 'application/json':
        return JsonResponse({'status': 'ok', 'grand_total': grand_total, 'items': cart})
    return redirect('/sales')
    

def scan_add_to_sale(request):
    """
    AJAX endpoint: receive 'isbn' (POST) from scanner and add product to sale_order list.
    Returns JSON with status and rendered HTML snippet for the order summary.
    """
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid method'}, status=405)

    isbn = request.POST.get('isbn', '').strip()
    if not isbn:
        return JsonResponse({'status': 'error', 'message': 'No ISBN provided'}, status=400)

    try:
        product = models.Product.objects.get(isbn=isbn)
    except Exception:
        return JsonResponse({'status': 'not_found', 'message': 'Product not found'})

    if product.quantity <= 0:
        return JsonResponse({'status': 'error', 'message': _('Cannot sell "%(product)s": out of stock.') % {'product': product.product_name}})

    # find existing session-backed sale_order list and increment or add
    cart = _get_sale_cart(request)
    for item in cart:
        if int(item.get('product_id')) == int(product.id):
            new_quantity = int(item.get('quantity', 0)) + 1
            if new_quantity > product.quantity:
                return JsonResponse({'status': 'error', 'message': _('Cannot sell more than available stock for "%(product)s".') % {'product': product.product_name}})
            item['quantity'] = new_quantity
            try:
                item['total_price'] = int(item['quantity']) * float(item.get('sale_price', 0))
            except Exception:
                item['total_price'] = 0
            break
    else:
        if 1 > product.quantity:
            return JsonResponse({'status': 'error', 'message': _('Cannot sell more than available stock for "%(product)s".') % {'product': product.product_name}})
        sale_price = float(product.sale_price) if product.sale_price is not None else 0.0
        cart.append({
            'product_name': product.product_name,
            'product_id': product.id,
            'quantity': 1,
            'sale_price': sale_price,
            'total_price': sale_price,
        })

    _save_sale_cart(request, cart)

    # compute grand total
    grand_total = sum(float(item.get('total_price', 0)) for item in cart)

    # return simplified JSON payload; front-end will re-render
    return JsonResponse({'status': 'ok', 'grand_total': grand_total, 'items': cart})

def submet_sale_order(request):
    cart = _get_sale_cart(request)
    if not cart:
        messages.error(request, _("Please add at least one product to sale!"))
        return redirect('/sales')
    else:
        employee_id = request.session['employee_id']
        # Get payment method from form POST
        pay_method = request.POST.get('invoice_pay_method', 'cash')
        # Read optional customer_id from the form
        customer_id = request.POST.get('customer_id') or None
        if customer_id:
            try:
                customer_id = int(customer_id)
            except Exception:
                customer_id = None
         
         
        
        
        # Get and validate currency
        currency_id = request.POST.get('currency_id')
        if not currency_id:
            messages.error(request, _("Please select a currency before submitting the sale!"))
            return redirect('/sales')
        
        
        # # product.currency
        # for pro_curr in cart:
            
        #     product_id = pro_curr.get('product_id')
            
        #     print(currency_id)
        #     product_currency = models.Product.objects.get(id=product_id).currency
        #     print(product_currency)
        #     # if product_currency and product_currency.id != int(currency_id):
        #     if product_currency and (product_currency.id != (models.CompanyProfile.objects.first().base_currency_id if models.CompanyProfile.objects.first() else None)):
                
        #         messages.error(request, _("اسف عملة المنتج لا تساوي عملة البيع!"))
        #         return redirect('/sales')
                
        
        # Create sale order and set payment method, customer, and currency
        sale_order = models.create_sale_order(employee_id, customer_id)
        sale_order.invoice_pay_method = pay_method
        
        # Set currency and calculate exchange rate if currency is selected
        if currency_id:
            try:
                sale_order.currency_id = currency_id
                # Get exchange rate for selected currency to base currency
                from datetime import date
                # Get exchange rate for selected currency ()
                
                # means: true if you defined an exchange_rate in your ExchangeRate TABLE in your company profile from currency to curency 
                # means: is there an ExchangeRate from (purchased currenct to BASE_currency) ?
                # ----EXAMBLE----
                # if you didnt add in (ExchangeRate TABLE) USD to USD :
                # if PURCHASE_CURRENCT == USD
                # and BASE_CURRENCY == USD
                # USD to USD : exchange_rate_obj == False
                # exchange_rate_obj = models.ExchangeRate.objects.filter(
                #     from_currency_id=currency_id,
                #     to_currency__id=models.CompanyProfile.objects.first().base_currency_id if models.CompanyProfile.objects.first() and models.CompanyProfile.objects.first().base_currency else None,
                #     date=date.today()
                # ).first()
                exchange_rate_obj = models.ExchangeRate.objects.filter(
                    from_currency_id=models.CompanyProfile.objects.first().base_currency_id if models.CompanyProfile.objects.first() and models.CompanyProfile.objects.first().base_currency else None,
                    to_currency__id=currency_id,
                    date=date.today()).first()
                
                # if found a record in the table (ExchangeRate)
                # git that rate 
                # put it in purchase.exchange_rate_to_base
                if exchange_rate_obj:
                    sale_order.exchange_rate_to_base = exchange_rate_obj.rate
                elif currency_id != (models.CompanyProfile.objects.first().base_currency_id if models.CompanyProfile.objects.first() else None):
                    # If no rate found and it's not the base currency, set rate to 1.0 as fallback
                    sale_order.exchange_rate_to_base = 1.0
            except Exception as e:
                messages.warning(request, _("Could not set exchange rate: %(err)s") % {'err': str(e)})

        sale_order.save()
        

        for key in cart:
            product_id = key.get('product_id')
            quantity = int(key.get('quantity'))
            sale_price = float(key.get('sale_price') or 0)
            total_price = float(key.get('total_price') or 0)
            models.add_item_to_invoice(product_id, quantity, sale_price, total_price)
            models.add_product_to_sale(product_id, quantity)

        # update the created Purchase total (grand_total and total_amount)
        try:
            
            last_sale = models.Sale_order.objects.last()
            total_sum = models.Sale_item.objects.filter(sale_order=last_sale).aggregate(total=Sum('total_price'))['total'] or 0
            last_sale.grand_total = total_sum
            last_sale.total_amount = total_sum            
            
            #_#_#_#_#_#_#_#_# (exchange_rate_to_base)
            # Calculate total_price_base if currency and exchange rate are set
            if last_sale.currency_id and last_sale.exchange_rate_to_base:
                last_sale.total_amount_base = total_sum * last_sale.exchange_rate_to_base
                # Also update (sale_items) with total_price_base
                for item in models.Sale_item.objects.filter(sale_order=last_sale):
                    item.currency_id = last_sale.currency_id
                    item.exchange_rate_to_base = last_sale.exchange_rate_to_base
                    item.total_price_base = item.total_price * last_sale.exchange_rate_to_base
                    item.save()
                    
            last_sale.save() 
        except Exception as e:
            pass

        _save_sale_cart(request, [])
        messages.success(request, _("Successfully Sold!"), extra_tags='sold_product')

        return redirect('/sales')
#____________________________________SALE___________________________________


def _get_sale_cart(request):
    return request.session.get('sale_cart', [])


def _save_sale_cart(request, cart):
    request.session['sale_cart'] = cart
    request.session.modified = True


def _get_purchase_cart(request):
    return request.session.get('purchase_cart', [])


def _save_purchase_cart(request, cart):
    request.session['purchase_cart'] = cart
    request.session.modified = True


def delete_product_from_sale(request):
    """
    AJAX endpoint: remove a product from the sale cart by product_id.
    Recalculates grand total and returns updated cart items.
    """
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid method'}, status=405)

    product_id = request.POST.get('product_id')
    if not product_id:
        return JsonResponse({'status': 'error', 'message': 'Product ID required'})

    cart = _get_sale_cart(request)
    # Remove the item with matching product_id
    cart = [item for item in cart if str(item.get('product_id')) != str(product_id)]
    _save_sale_cart(request, cart)

    # Recalculate grand total
    grand_total = sum(float(item.get('total_price', 0)) for item in cart)

    return JsonResponse({'status': 'ok', 'grand_total': grand_total, 'items': cart})

def update_sales_cart(request):
    """
    Persist edited sales cart rows to session.
    Accepts JSON payload: { items: [ {product_id, quantity, sale_price} ] }
    Returns updated items and grand_total.
    """
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid method'}, status=405)

    items_payload = None
    if request.content_type == 'application/json':
        import json
        try:
            body = json.loads(request.body or '{}')
            items_payload = body.get('items')
        except Exception:
            return JsonResponse({'status': 'error', 'message': _('Invalid payload')}, status=400)
    else:
        # Fallback: accept form-encoded 'items' JSON string
        items_str = request.POST.get('items')
        if items_str:
            import json
            try:
                items_payload = json.loads(items_str)
            except Exception:
                items_payload = None

    if not isinstance(items_payload, list):
        return JsonResponse({'status': 'error', 'message': _('No items provided')}, status=400)

    new_cart = []
    for it in items_payload:
        product_id = it.get('product_id')
        quantity = it.get('quantity')
        sale_price = it.get('sale_price')
        try:
            product_id = int(product_id)
        except (ValueError, TypeError):
            return JsonResponse({'status': 'error', 'message': _('Invalid product_id')}, status=400)
        try:
            quantity = int(quantity)
        except (ValueError, TypeError):
            quantity = 0
        try:
            sale_price = float(sale_price)
        except (ValueError, TypeError):
            sale_price = 0.0

        if quantity <= 0:
            return JsonResponse({'status': 'error', 'message': _('Quantity must be positive')}, status=400)

        # Validate against available stock
        try:
            prod = models.Product.objects.get(id=product_id)
        except models.Product.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': _('Product not found')}, status=404)

        available_qty = int(prod.quantity or 0)
        if available_qty <= 0:
            msg = _('Cannot sell "%(product)s": out of stock.') % {'product': prod.product_name}
            return JsonResponse({'status': 'error', 'message': msg}, status=400)
        if quantity > available_qty:
            msg = _('Cannot sell more than available stock for "%(product)s". Available: %(available)d, requested: %(requested)d') % {
                'product': prod.product_name, 'available': available_qty, 'requested': quantity
            }
            return JsonResponse({'status': 'error', 'message': msg}, status=400)

        total_price = quantity * sale_price
        new_cart.append({
            'product_name': prod.product_name,
            'product_id': product_id,
            'quantity': quantity,
            'sale_price': sale_price,
            'total_price': total_price,
        })

    _save_sale_cart(request, new_cart)

    grand_total = sum(float(item.get('total_price', 0)) for item in new_cart)
    return JsonResponse({'status': 'ok', 'grand_total': grand_total, 'items': new_cart})


def delete_product_from_purchase(request):
    """
    AJAX endpoint: remove a product from the purchase cart by product_id.
    Recalculates grand total and returns updated cart items.
    """
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid method'}, status=405)

    product_id = request.POST.get('product_id')
    if not product_id:
        return JsonResponse({'status': 'error', 'message': 'Product ID required'})

    cart = _get_purchase_cart(request)
    # Remove the item with matching product_id
    cart = [item for item in cart if str(item.get('product_id')) != str(product_id)]
    _save_purchase_cart(request, cart)

    # Recalculate grand total
    grand_total = sum(float(item.get('total_price', 0)) for item in cart)

    return JsonResponse({'status': 'ok', 'grand_total': grand_total, 'items': cart})


# -------------------- Sales Return (customer returns) --------------------
def sale_return_create_view(request, sale_order_id):
    # ensure employee logged in
    if 'employee_id' not in request.session:
        messages.error(request, _('Please sign in to create a return.'))
        return redirect('/')

    try:
        sale_order = models.Sale_order.objects.get(id=sale_order_id)
    except models.Sale_order.DoesNotExist:
        messages.error(request, _('Sale order not found.'))
        return redirect('/sales')

    # build items with allowed return quantities (per original sale_item)
    items = []
    for si in sale_order.sale_items.all():
        returned_qty = models.SaleReturnItem.objects.filter(original_item=si).aggregate(total=Sum('quantity'))['total'] or 0
        allowed = max(0, si.quantity - int(returned_qty))
        items.append({
            'sale_item': si,
            'product': si.product,
            'sold_quantity': si.quantity,
            'unit_price': si.unit_price,
            'allowed_return_quantity': allowed,
        })

    if request.method == 'POST':
        # collect requested returns
        requested = []
        for entry in items:
            si = entry['sale_item']
            key = f'return_qty_{si.id}'
            qty_str = request.POST.get(key, '0')
            try:
                qty = int(qty_str)
            except ValueError:
                qty = 0
            if qty > 0:
                if qty > entry['allowed_return_quantity']:
                    messages.error(request, _('Cannot return %(qty)d of %(product)s; only %(allowed)d available to return.') % {'qty': qty, 'product': si.product.product_name, 'allowed': entry['allowed_return_quantity']})
                    return redirect(request.path)
                requested.append((si, qty))

        if not requested:
            messages.error(request, _('Please enter at least one quantity to return.'))
            return redirect(request.path)

        # create SaleReturn and items
        employee_id = request.session['employee_id']
        # use Django transaction when available
        # to ensure ALL - OR - NOTHING OPERATION
        with transaction.atomic():
            # use helper to create return (keeps consistency with existing code)
            models.create_sale_return(employee_id, sale_order_id)
            for si, qty in requested:
                unit_price = float(si.unit_price or 0)
                total_price = qty * unit_price
                models.add_item_to_sale_return(si.product.id, qty, unit_price, total_price, original_item_id=si.id)
                # increase stock on return
                models.add_product_back_on_return(si.product.id, qty)

            # update totals on the created SaleReturn
            try:
                grand = sum(qty * float(si.unit_price or 0) for si, qty in requested)
            except Exception:
                grand = 0.0
            last_sr = models.SaleReturn.objects.last()
            if last_sr:
                last_sr.grand_total = grand
                last_sr.total_amount = grand
                last_sr.save()

        messages.success(request, _('Sale return recorded.'), extra_tags='sale_return')
        return redirect(f'/sales/returns/{ last_sr.id }')

    context = {
        'sale_order': sale_order,
        'items': items,
        'employee': models.get_employee_by_id(request.session['employee_id']) if 'employee_id' in request.session else None,
    }
    return render(request, 'sales/sale_return_form.html', context)


def sale_return_detail_view(request, id):
    try:
        sr = models.SaleReturn.objects.get(id=id)
    except models.SaleReturn.DoesNotExist:
        messages.error(request, _('Sale return not found.'))
        return redirect('/sales')

    items = sr.items.select_related('product', 'original_item').all()
    total = items.aggregate(total=Sum('total_price'))['total'] or 0
    context = {
        'sale_return': sr,
        'items': items,
        'total': total,
    }
    return render(request, 'sales/sale_return_detail.html', context)


# -------------------- Purchase Return (supplier returns) --------------------
def purchase_return_create_view(request, purchase_id):
    if 'employee_id' not in request.session:
        messages.error(request, _('Please sign in to create a return.'))
        return redirect('/')

    try:
        purchase = models.Purchase.objects.get(id=purchase_id)
    except models.Purchase.DoesNotExist:
        messages.error(request, _('Purchase not found.'))
        return redirect('/purchases')

    items = []
    for pi in purchase.purchase_items.all():
        returned_qty = models.Return_item.objects.filter(original_item=pi).aggregate(total=Sum('quantity'))['total'] or 0
        allowed = max(0, pi.quantity - int(returned_qty))
        items.append({
            'purchase_item': pi,
            'product': pi.product,
            'purchased_quantity': pi.quantity,
            'unit_price': pi.unit_price,
            'allowed_return_quantity': allowed,
        })

    if request.method == 'POST':
        requested = []
        for entry in items:
            pi = entry['purchase_item']
            key = f'return_qty_{pi.id}'
            qty_str = request.POST.get(key, '0')
            try:
                qty = int(qty_str)
            except ValueError:
                qty = 0
            if qty > 0:
                if qty > entry['allowed_return_quantity']:
                    messages.error(request, _('Cannot return %(qty)d of %(product)s; only %(allowed)d available to return.') % {'qty': qty, 'product': pi.product.product_name, 'allowed': entry['allowed_return_quantity']})
                    return redirect(request.path)
                requested.append((pi, qty))

        if not requested:
            messages.error(request, _('Please enter at least one quantity to return.'))
            return redirect(request.path)

        employee_id = request.session['employee_id']
        with transaction.atomic():
            models.create_return_order(employee_id)
            for pi, qty in requested:
                unit_price = float(pi.unit_price or 0)
                total_price = qty * unit_price
                models.add_item_to_return_invoice(pi.product.id, qty, unit_price, total_price, original_item_id=pi.id)
                # For purchase returns, decrease stock because products are returned to supplier
                models.add_product_to_return(pi.product.id, qty)

            try:
                grand = sum(qty * float(pi.unit_price or 0) for pi, qty in requested)
            except Exception:
                grand = 0.0
            last_r = models.Return.objects.last()
            if last_r:
                last_r.grand_total = grand
                last_r.total_amount = grand
                last_r.save()

        messages.success(request, _('Purchase return recorded.'), extra_tags='return_purchase')
        return redirect(f'/purchases/returns/{ last_r.id }')

    context = {
        'purchase': purchase,
        'items': items,
        'employee': models.get_employee_by_id(request.session['employee_id']) if 'employee_id' in request.session else None,
    }
    return render(request, 'purchases/purchase_return_form.html', context)


def purchase_return_detail_view(request, id):
    try:
        r = models.Return.objects.get(id=id)
    except models.Return.DoesNotExist:
        messages.error(request, _('Return not found.'))
        return redirect('/purchases')

    items = r.return_items.select_related('product', 'original_item').all()
    total = items.aggregate(total=Sum('total_price'))['total'] or 0
    context = {
        'return': r,
        'items': items,
        'total': total,
    }
    return render(request, 'purchases/purchase_return_detail.html', context)



# session-backed purchases_order is stored per-user in request.session via helpers above
#____________________________________PURCHASE___________________________________
def add_product_to_purchase(request):
    # Support both normal POST and AJAX/JSON requests. When called via AJAX return JSON
    data = None
    if request.content_type == 'application/json':
        import json
        try:
            data = json.loads(request.body)
        except Exception:
            return JsonResponse({'status': 'error', 'message': _('Invalid payload')}, status=400)

    # extract fields from either JSON payload or form POST
    product_name = (data.get('product_name') if data else request.POST.get('product_name'))
    quantity = int((data.get('quantity') if data else request.POST.get('quantity', 0)) or 0)

    if not product_name or quantity <= 0:
        msg = _('Product name and positive quantity are required.')
        if request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.content_type == 'application/json':
            return JsonResponse({'status': 'error', 'message': msg}, status=400)
        messages.error(request, msg)
        return redirect('/purchases')

    try:
        prod = models.Product.objects.get(product_name=product_name)
    except models.Product.DoesNotExist:
        msg = _('Product not found')
        if request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.content_type == 'application/json':
            return JsonResponse({'status': 'not_found', 'message': msg})
        messages.error(request, msg)
        return redirect('/purchases')

    product_id = prod.id
    purchase_price = float(prod.purchasing_price) if prod.purchasing_price is not None else 0.0
    total_price = quantity * purchase_price
    cart = _get_purchase_cart(request)
    # if product already in cart, increase quantity
    for item in cart:
        if int(item.get('product_id')) == int(product_id):
            item['quantity'] = int(item.get('quantity', 0)) + quantity
            try:
                item['total_price'] = int(item['quantity']) * float(item.get('purchase_price', 0))
            except Exception:
                item['total_price'] = 0
            break
    else:
        cart.append({
            'product_name': product_name,
            'product_id': product_id,
            'quantity': quantity,
            'purchase_price': purchase_price,
            'total_price': total_price,
        })
    _save_purchase_cart(request, cart)

    # compute grand total
    grand_total = sum(float(item.get('total_price', 0)) for item in cart)
    if request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.content_type == 'application/json':
        return JsonResponse({'status': 'ok', 'grand_total': grand_total, 'items': cart})
    return redirect('/purchases')
    

def scan_add_to_purchase(request):
    """
    AJAX endpoint: receive 'isbn' (POST) from scanner and add product to purchases_order list.
    """
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid method'}, status=405)

    # Support AJAX JSON payload
    try:
        if request.content_type == 'application/json':
            import json
            data = json.loads(request.body)
            isbn = data.get('isbn', '').strip()
        else:
            isbn = request.POST.get('isbn', '').strip()
    except Exception:
        return JsonResponse({'status': 'error', 'message': 'Invalid payload'}, status=400)

    if not isbn:
        return JsonResponse({'status': 'error', 'message': 'No ISBN provided'}, status=400)

    try:
        product = models.Product.objects.get(isbn=isbn)
    except Exception:
        return JsonResponse({'status': 'not_found', 'message': 'Product not found'})

    cart = _get_purchase_cart(request)
    for item in cart:
        if int(item.get('product_id')) == int(product.id):
            item['quantity'] = int(item.get('quantity', 0)) + 1
            try:
                item['total_price'] = int(item['quantity']) * float(item.get('purchase_price', 0))
            except Exception:
                item['total_price'] = 0
            break
    else:
        purchase_price = float(product.purchasing_price) if product.purchasing_price is not None else 0.0
        cart.append({
            'product_name': product.product_name,
            'product_id': product.id,
            'quantity': 1,
            'purchase_price': purchase_price,
            'total_price': purchase_price,
        })

    _save_purchase_cart(request, cart)

    # compute grand total from session cart
    cart = _get_purchase_cart(request)
    grand_total = sum(float(item.get('total_price', 0)) for item in cart)
    return JsonResponse({'status': 'ok', 'grand_total': grand_total, 'items': cart})

#_#_#_#_#_#_#_#_#_#_#_#_#_#_#_#_# --- Puchase ---_#_#_#_#_#_#_#_#_#_#_#_#_#_#
def submet_purchase_order(request):
    cart = _get_purchase_cart(request)
    if not cart:
        messages.error(request, _("Please add at least one product to purchase!"))
        return redirect('/purchases')
    
    # Validate currency selection
    currency_id = request.POST.get('currency_id')
    if not currency_id:
        messages.error(request, _("Please select a currency before submitting the purchase!"))
        return redirect('/purchases')
    
    employee_id = request.session['employee_id']
    # Get payment method from form POST
    pay_method = request.POST.get('invoice_pay_method', 'cash')
    supplier_id = request.POST.get('supplier_id')
    
    # create purchase order with optional supplier
    # sala_order for sales
    purchase = models.create_purchase_order(employee_id, supplier_id=supplier_id if supplier_id else None)
    purchase.invoice_pay_method = pay_method
    
    # Set currency and calculate exchange rate if currency is selected
    if currency_id:
        try:
            purchase.currency_id = currency_id
            # Get exchange rate for selected currency to base currency
            from datetime import date
            # Get exchange rate for selected currency ()
            
            # means: true if you defined an exchange_rate in your ExchangeRate TABLE in your company profile from currency to curency 
            # means: is there an ExchangeRate from (purchased currenct to BASE_currency) ?
            # ----EXAMBLE----
            # if you didnt add in (ExchangeRate TABLE) USD to USD :
            # if PURCHASE_CURRENCT == USD
            # and BASE_CURRENCY == USD
            # USD to USD : exchange_rate_obj == False
            exchange_rate_obj = models.ExchangeRate.objects.filter(
                from_currency_id=currency_id,
                to_currency__id=models.CompanyProfile.objects.first().base_currency_id if models.CompanyProfile.objects.first() and models.CompanyProfile.objects.first().base_currency else None,
                date=date.today()
            ).first()
            
            # if found a record in the table (ExchangeRate)
            # git that rate 
            # put it in purchase.exchange_rate_to_base
            if exchange_rate_obj:
                purchase.exchange_rate_to_base = exchange_rate_obj.rate
            elif currency_id != (models.CompanyProfile.objects.first().base_currency_id if models.CompanyProfile.objects.first() else None):
                # If no rate found and it's not the base currency, set rate to 1.0 as fallback
                purchase.exchange_rate_to_base = 1.0
        except Exception as e:
            messages.warning(request, _("Could not set exchange rate: %(err)s") % {'err': str(e)})
    
    purchase.save()
    
    for key in cart:
        product_id = key.get('product_id')
        quantity = int(key.get('quantity'))
        purchase_price = float(key.get('purchase_price') or 0)
        total_price = float(key.get('total_price') or 0)
        models.add_item_to_purchase_invoice(product_id, quantity, purchase_price, total_price)
        models.add_product_to_purchase(product_id, quantity)

    # update the created Purchase total (grand_total and total_amount)
    try:
        last_purchase = models.Purchase.objects.last()
        total_sum = models.Purchase_item.objects.filter(purchase_id=last_purchase).aggregate(total=Sum('total_price'))['total'] or 0
        last_purchase.grand_total = total_sum
        last_purchase.total_amount = total_sum
        
        #_#_#_#_#_#_#_#_# (exchange_rate_to_base)
        # Calculate total_price_base if currency and exchange rate are set
        if last_purchase.currency_id and last_purchase.exchange_rate_to_base:
            last_purchase.total_price_base = total_sum * last_purchase.exchange_rate_to_base
            # Also update (purchase_items) with total_price_base
            for item in models.Purchase_item.objects.filter(purchase_id=last_purchase):
                item.currency_id = last_purchase.currency_id
                item.exchange_rate_to_base = last_purchase.exchange_rate_to_base
                item.total_price_base = item.total_price * last_purchase.exchange_rate_to_base
                item.save()
                
                # Update product purchasing_price if currency is not base currency
                # If purchased currency differs from base currency, update product purchasing_price
                # (عملة الشركة)يتم تحديث سعر الشراء للمنتج إذا كانت عملة الشراء مختلفة عن العملة الأساسية
                base_currency_id = models.CompanyProfile.objects.first().base_currency_id if models.CompanyProfile.objects.first() else None
                if last_purchase.currency_id != base_currency_id:
                    try:
                        product = models.Product.objects.get(id=item.product_id)
                        # Set purchasing_price to the price in base currency
                        # product.purchasing_price = float(item.purchase_price) * float(last_purchase.exchange_rate_to_base)
                        product.purchasing_price = float(item.unit_price) * float(last_purchase.exchange_rate_to_base)
                        product.save()
                        messages.warning(request, _("Updated purchasing price for product '%(prod)s' to %(price).2f in base currency.") % {'prod': product.product_name, 'price': product.purchasing_price})
                    except Exception as e:
                        # Optionally log or handle error
                        messages.warning(request, _("Could not Convert : %(err)s") % {'err': str(e)})
                        # pass
                        
        last_purchase.save()
    except Exception as e:
        messages.warning(request, _("Could not update purchase totals: %(err)s") % {'err': str(e)})

    _save_purchase_cart(request, [])
    messages.success(request, _("Purchased Successfully!"), extra_tags = 'add_invoice')

    return redirect('/purchases')
#____________________________________PURCHASE___________________________________

def clear_purchases_list(request):
    cart = _get_purchase_cart(request)
    if not cart:
        messages.error(request, _("already empty!"))
        return redirect('/purchases')
    _save_purchase_cart(request, [])
    return redirect('/purchases')


# -------------------- RETURN PURCHASES (Products returned) --------------------
returns_order = []

def display_returns(request):
    # calculate grand total for current return order
    grand_total = sum(item['total_price'] for item in returns_order)

    # Paginate recent returned invoices
    return_invoices_qs = models.get_all_return_invoices()
    return_invoices_page_number = request.GET.get('return_invoices_page')
    return_invoices_paginator = Paginator(return_invoices_qs, 10)  # 10 items per page
    return_invoices_page = return_invoices_paginator.get_page(return_invoices_page_number)

    context = {
        'returns_order': returns_order,
        'products': models.get_all_products(),
        'return_invoices': return_invoices_page,  # Updated to use paginated data
        'employee': models.get_employee_by_id(request.session['employee_id']),
        'grand_total': grand_total,
        'return_invoices_paginator': return_invoices_paginator,
    }
    return render(request, 'return_purchases.html', context)


def add_product_to_return(request):
    errors = models.Purchase.objects.invoice_validator(request.POST)
    if len(errors) > 0:
        for key, value in errors.items():
            messages.error(request, _(value))
        return redirect('/return_purchases')
    else:
        product_name = request.POST['product_name']
        quantity = request.POST['quantity']
        product = models.Product.objects.get(product_name=product_name)
        product_id = product.id
        unit_price = product.purchasing_price
        total_price = int(quantity) * float(unit_price)
        returns_order.append({
            'product_name': product_name,
            'product_id': product_id,
            'quantity': quantity,
            'unit_price': unit_price,
            'total_price': total_price,
        })
        # immediately subtract quantity from product stock in memory via model helper
        models.add_product_to_return(product_id, quantity)
        return redirect('/return_purchases')


def submet_return_order(request):
    if returns_order == []:
        messages.error(request, _("Please add at least one product to return!"))
        return redirect('/return_purchases')
    else:
        employee_id = request.session['employee_id']
        # Get payment method from form POST
        pay_method = request.POST.get('invoice_pay_method', 'cash')
        ret = models.create_return_order(employee_id)
        ret.invoice_pay_method = pay_method
        ret.save()
        for key in returns_order:
            product_id = key.get('product_id')
            quantity = key.get('quantity')
            unit_price = key.get('unit_price')
            total_price = key.get('total_price')
            models.add_item_to_return_invoice(product_id, quantity, unit_price, total_price)
            # Already subtracted when adding to cart, so no need to call again

        # compute grand_total from the items we've just added
        try:
            grand = sum(float(item.get('total_price')) for item in returns_order)
        except Exception:
            grand = 0.0
        # update grand_total on the created return invoice
        last_return = models.Return.objects.last()
        if last_return:
            last_return.grand_total = grand
            last_return.total_amount = grand
            last_return.save()

        returns_order.clear()
        messages.success(request, _("Return Recorded Successfully!"), extra_tags='add_return')
        return redirect('/return_purchases')


def clear_returns_list(request):
    if returns_order == []:
        messages.error(request, _("already empty!"))
        return redirect('/return_purchases')
    else:
        # If user clears the list, we should restore quantities back to stock
        for item in returns_order:
            try:
                models.add_product_to_purchase(item.get('product_id'), item.get('quantity'))
            except Exception:
                pass
        returns_order.clear()
        return redirect('/return_purchases')


def view_return_invoice(request, id):
    context = {
        'invoice': models.get_return_invoice(id),
        'return_products': models.return_invoices_products(id),
    }
    return render(request, 'view_return_invoice.html', context)


# ---------- Sale Returns (customer returns) session-backed cart

def _get_sale_returns_cart(request):
    return request.session.get('sale_returns_cart', [])


def _save_sale_returns_cart(request, cart):
    request.session['sale_returns_cart'] = cart
    request.session.modified = True


def display_sale_returns(request):
    cart = _get_sale_returns_cart(request)
    # ensure numeric floats for display
    grand_total = sum(float(item.get('total_price', 0)) for item in cart)

    # Paginate recent sale returns
    sale_returns_qs = models.get_all_sale_returns()
    sale_returns_page_number = request.GET.get('sale_returns_page')
    sale_returns_paginator = Paginator(sale_returns_qs, 10)  # 10 items per page
    sale_returns_page = sale_returns_paginator.get_page(sale_returns_page_number)

    context = {
        'sale_returns_cart': cart,
        'products': models.get_all_products(),
        'sale_returns': sale_returns_page,  # Updated to use paginated data
        'employee': models.get_employee_by_id(request.session['employee_id']),
        'grand_total': grand_total,
        'sale_returns_paginator': sale_returns_paginator,
    }
    return render(request, 'return_sales.html', context)


def add_product_to_sale_return(request):
    errors = models.Purchase.objects.invoice_validator(request.POST)
    if len(errors) > 0:
        for key, value in errors.items():
            messages.error(request, value)
        return redirect('/return_sales')
    else:
        product_name = request.POST['product_name']
        quantity = int(request.POST['quantity'])
        product = models.Product.objects.get(product_name=product_name)
        product_id = product.id
        unit_price = float(product.sale_price) if product.sale_price is not None else 0.0
        total_price = quantity * float(unit_price)

        cart = _get_sale_returns_cart(request)
        cart.append({
            'product_name': product_name,
            'product_id': product_id,
            'quantity': quantity,
            'unit_price': float(unit_price),
            'total_price': float(total_price),
        })
        _save_sale_returns_cart(request, cart)

        # restock the product immediately
        models.add_product_back_on_return(product_id, quantity)
        return redirect('/return_sales')


def submit_sale_return(request):
    cart = _get_sale_returns_cart(request)
    if not cart:
        messages.error(request, "Please add at least one product to return!")
        return redirect('/return_sales')
    else:
        from django.db import transaction
        employee_id = request.session['employee_id']
        sale_order_id = request.POST.get('sale_order_id') if request.method == 'POST' else None

        # If a sale_order_id was chosen, validate each cart item belongs to that sale and quantity <= allowed (sold - already returned)
        if sale_order_id:
            try:
                so = models.Sale_order.objects.get(id=sale_order_id)
            except Exception:
                messages.error(request, "Selected sale order not found.")
                return redirect('/return_sales')

            # build sold map and already-returned map
            sold_qs = models.Sale_item.objects.filter(sale_order=so).values('product_id').annotate(total=Sum('quantity'))
            sold_map = {row['product_id']: int(row['total'] or 0) for row in sold_qs}
            returned_qs = models.SaleReturnItem.objects.filter(sale_return__sale_order=so).values('product_id').annotate(total=Sum('quantity'))
            returned_map = {row['product_id']: int(row['total'] or 0) for row in returned_qs}

            for item in cart:
                pid = item.get('product_id')
                requested = int(item.get('quantity'))
                sold = sold_map.get(pid, 0)
                already = returned_map.get(pid, 0)
                allowed = max(0, sold - already)
                if sold == 0:
                    messages.error(request, f"Product id {pid} is not part of sale invoice #{sale_order_id}.")
                    return redirect('/return_sales')
                if requested > allowed:
                    messages.error(request, f"Cannot return {requested} of product id {pid}; only {allowed} remaining from invoice #{sale_order_id}.")
                    return redirect('/return_sales')

        # Passed validations; create SaleReturn and items inside a transaction, linking to original sale_item when possible
        with transaction.atomic():
            # Get payment method from form POST
            pay_method = request.POST.get('invoice_pay_method', 'cash')
            sr = models.create_sale_return(employee_id, sale_order_id)
            sr.invoice_pay_method = pay_method
            sr.save()
            for key in cart:
                product_id = key.get('product_id')
                quantity = int(key.get('quantity'))
                unit_price = float(key.get('unit_price'))
                total_price = float(key.get('total_price'))
                original_item_id = None
                if sale_order_id:
                    orig = models.Sale_item.objects.filter(sale_order_id=sale_order_id, product_id=product_id).first()
                    if orig:
                        original_item_id = orig.id
                models.add_item_to_sale_return(product_id, quantity, unit_price, total_price, original_item_id)

            # compute grand total and save to sale return record
            try:
                grand = sum(float(item.get('total_price')) for item in cart)
            except Exception:
                grand = 0.0
            last_sr = models.SaleReturn.objects.last()
            if last_sr:
                last_sr.grand_total = grand
                last_sr.total_amount = grand
                last_sr.save()

        # clear session cart and return
        _save_sale_returns_cart(request, [])
        messages.success(request, "Sale return recorded!", extra_tags='sale_return')
        return redirect('/return_sales')


def clear_sale_returns_cart(request):
    cart = _get_sale_returns_cart(request)
    if not cart:
        messages.error(request, "already empty!")
        return redirect('/return_sales')
    else:
        # if clearing, reverse the earlier restock
        for item in cart:
            try:
                models.add_product_to_sale(item.get('product_id'), int(item.get('quantity')))
            except Exception:
                pass
        _save_sale_returns_cart(request, [])
        return redirect('/return_sales')


def view_sale_return(request, id):
    context = {
        'invoice': models.get_sale_return(id),
        'return_products': models.sale_return_products(id),
        'company': models.get_company_profile(),
    }
    return render(request, 'view_return_sale.html', context)


def print_sale_return(request, id):
    context = {
        'invoice': models.get_sale_return(id),
        'return_products': models.sale_return_products(id),
        'company': models.get_company_profile(),
    }
    return render(request, 'print_return_sale.html', context)


def print_return_invoice(request, id):
    # standalone printable page for the return invoice
    context = {
        'invoice': models.get_return_invoice(id),
        'return_products': models.return_invoices_products(id),
        'company': models.get_company_profile(),
    }
    return render(request, 'print_return_invoice.html', context)


def print_sale_invoice(request, id):
    context = {
        'order': models.get_sale_order(id),
        'sale_products': models.sale_orders_products(id),
        'company': models.get_company_profile(),

    }
    return render(request, 'print_sale_invoice.html', context)


def print_purchase_invoice(request, id):
    context = {
        'invoice': models.get_purchases(id),
        'purchase_products': models.purchase_invoices_products(id),
        'company': models.get_company_profile(),
    }
    return render(request, 'print_purchase_invoice.html', context)


def print_stock_out_voucher(request, id):
    try:
        voucher = models.Stock_Out_Voucher.objects.get(id=id)
    except models.Stock_Out_Voucher.DoesNotExist:
        messages.error(request, _('Voucher not found.'))
        return redirect('/stock_out_voucher')

    items = voucher.Stock_Out_Voucher_items.select_related('product').all()
    context = {
        'voucher': voucher,
        'items': items,
        'company': models.get_company_profile(),
    }
    return render(request, 'print_stock_out_voucher.html', context)


def clear_sales_list(request) :
    cart = _get_sale_cart(request)
    if not cart:
        messages.error(request, "already empty!")
        return redirect('/sales')
    _save_sale_cart(request, [])
    return redirect('/sales')

def display_employee_reports(request):
    context = {
        'products': models.get_all_products(),
        'today': datetime.date.today(),
        'expiry_range': datetime.date.today() + timedelta(days=6*30),
        'employee': models.get_employee_by_id(request.session['employee_id'])
    }
    return render (request, 'employee_reports.html', context)


def display_company_profile(request):
    # Only allow logged-in employees to edit company profile
    if 'employee_id' not in request.session:
        return redirect('/index')

    if request.method == 'POST':
        # Handle company profile update
        if 'save_company' in request.POST:
            company_name = request.POST.get('company_name', '').strip()
            registration_number = request.POST.get('registration_number', '').strip()
            address = request.POST.get('address', '').strip()
            phone = request.POST.get('phone', '').strip()
            email = request.POST.get('email', '').strip()
            base_currency_id = request.POST.get('base_currency')

            if not company_name:
                messages.error(request, _('Company name is required.'))
                return redirect('/company_profile')

            try:
                base_currency = None
                if base_currency_id:
                    try:
                        base_currency = models.Currency.objects.get(id=base_currency_id)
                    except models.Currency.DoesNotExist:
                        pass
                
                company = models.get_company_profile()
                if company:
                    company.company_name = company_name
                    company.registration_number = registration_number or None
                    company.address = address or None
                    company.phone = phone or None
                    company.email = email or None
                    company.base_currency = base_currency
                    company.save()
                else:
                    company = models.CompanyProfile.objects.create(
                        company_name=company_name,
                        registration_number=registration_number or None,
                        address=address or None,
                        phone=phone or None,
                        email=email or None,
                        base_currency=base_currency,
                    )
                messages.success(request, _('Company profile saved.'))
            except Exception as e:
                messages.error(request, _('Failed to save company profile: %(err)s') % {'err': str(e)})

        # Handle new currency creation
        elif 'add_currency' in request.POST:
            code = request.POST.get('currency_code', '').strip().upper()
            name = request.POST.get('currency_name', '').strip()

            if not code or not name:
                messages.error(request, _('Currency code and name are required.'))
                return redirect('/company_profile')

            try:
                if models.Currency.objects.filter(code=code).exists():
                    messages.error(request, _('Currency with code %(code)s already exists.') % {'code': code})
                else:
                    models.Currency.objects.create(code=code, name=name)
                    messages.success(request, _('Currency added successfully.'))
            except Exception as e:
                messages.error(request, _('Failed to add currency: %(err)s') % {'err': str(e)})

        # Handle new exchange rate creation
        elif 'add_exchange_rate' in request.POST:
            from_currency_id = request.POST.get('from_currency')
            to_currency_id = request.POST.get('to_currency')
            rate = request.POST.get('exchange_rate', '')

            if not from_currency_id or not to_currency_id or not rate:
                messages.error(request, _('All exchange rate fields are required.'))
                return redirect('/company_profile')

            try:
                from_currency = models.Currency.objects.get(id=from_currency_id)
                to_currency = models.Currency.objects.get(id=to_currency_id)
                rate_value = float(rate)
                
                if rate_value <= 0:
                    messages.error(request, _('Exchange rate must be greater than 0.'))
                    return redirect('/company_profile')

                # Update or create exchange rate for today
                import datetime
                today = datetime.date.today()
                exchange_rate, created = models.ExchangeRate.objects.update_or_create(
                    from_currency=from_currency,
                    to_currency=to_currency,
                    date=today,
                    defaults={'rate': rate_value}
                )
                
                if created:
                    messages.success(request, _('Exchange rate added successfully.'))
                else:
                    messages.success(request, _('Exchange rate updated successfully.'))
            except models.Currency.DoesNotExist:
                messages.error(request, _('One or both currencies not found.'))
            except ValueError:
                messages.error(request, _('Exchange rate must be a valid number.'))
            except Exception as e:
                messages.error(request, _('Failed to add exchange rate: %(err)s') % {'err': str(e)})

        # Handle exchange rate edit
        elif 'edit_exchange_rate' in request.POST:
            rate_id = request.POST.get('rate_id')
            rate_value = request.POST.get('exchange_rate', '')

            if not rate_id or not rate_value:
                messages.error(request, _('Rate ID and value are required.'))
                return redirect('/company_profile')

            try:
                exchange_rate = models.ExchangeRate.objects.get(id=rate_id)
                rate_float = float(rate_value)
                
                if rate_float <= 0:
                    messages.error(request, _('Exchange rate must be greater than 0.'))
                    return redirect('/company_profile')

                exchange_rate.rate = rate_float
                exchange_rate.save()
                messages.success(request, _('Exchange rate updated successfully.'))
            except models.ExchangeRate.DoesNotExist:
                messages.error(request, _('Exchange rate not found.'))
            except ValueError:
                messages.error(request, _('Exchange rate must be a valid number.'))
            except Exception as e:
                messages.error(request, _('Failed to update exchange rate: %(err)s') % {'err': str(e)})

        # Handle exchange rate delete
        elif 'delete_exchange_rate' in request.POST:
            rate_id = request.POST.get('delete_exchange_rate')

            if not rate_id:
                messages.error(request, _('Rate ID is required.'))
                return redirect('/company_profile')

            try:
                exchange_rate = models.ExchangeRate.objects.get(id=rate_id)
                exchange_rate.delete()
                messages.success(request, _('Exchange rate deleted successfully.'))
            except models.ExchangeRate.DoesNotExist:
                messages.error(request, _('Exchange rate not found.'))
            except Exception as e:
                messages.error(request, _('Failed to delete exchange rate: %(err)s') % {'err': str(e)})

        return redirect('/company_profile')

    company = models.get_company_profile()
    currencies = models.Currency.objects.all()
    exchange_rates = models.ExchangeRate.objects.select_related('from_currency', 'to_currency').order_by('-date')[:10]
    
    context = {
        'company': company,
        'employee': models.get_employee_by_id(request.session['employee_id']),
        'currencies': currencies,
        'exchange_rates': exchange_rates,
    }
    return render(request, 'company_profile.html', context)

def view_sale_order(request, id):#--------------------------------------------Mai
    sale_qs = models.sale_orders_products(id)
    context={
        'order': models.get_sale_order(id),
        'sale_products': sale_qs,
        'sale_products_count': sale_qs.count(),
        'page_title': _("View Sale Order"),
        'no_products_message': _("No products found in this sale order."),
        'company': models.get_company_profile(),

    }
    return render(request, 'view_sale_order.html',context)

def view_purchase_invoice(request,id):
    context={
        # 'order': models.get_sale_order(id),
        'invoice': models.get_purchases(id),
        'purchase_products':models.purchase_invoices_products(id),#MAI*****
        'company': models.get_company_profile(),
    }
    return render(request, 'view_purchase_invoice.html',context)

def display_edit_form(request,id):
    context={
        'products': models.get_all_products(),
        'product':models.get_product(id),
        'employee': models.get_employee_by_id(request.session['employee_id'])
    }
    return render(request, 'edit_product.html',context)

def update_product(request,id):
    errors = models.Product.objects.product_validator(request.POST)
    if len(errors) > 0:
        for key, value in errors.items():
            messages.error(request, value)
        return redirect(f'/edit_product/{id}')
    else:
        models.update_selected_product(request,id)

    return redirect('/employye_dashboard')#--------------------------------------------Mai


def get_date_time():
    return datetime.date.today()# used in models line 69

from django.views.decorators.csrf import csrf_exempt

@csrf_exempt   # only if you want to test without CSRF
def search_results(request):
    if 'employee_id' not in request.session:
        return redirect('/index')

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        searchValue = request.POST.get('searchValue', '')
        qs = models.Product.objects.filter(product_name__icontains=searchValue)

        if qs.exists() and len(searchValue) > 0:
            data = []
            for pos in qs:
                data.append({
                    'id': pos.id,
                    'product_name': pos.product_name,
                    'quantity': pos.quantity,
                    'purchasing_price': str(pos.purchasing_price),  # safe for JSON
                    'expiry_date': pos.expiry_date.strftime('%Y-%m-%d') if pos.expiry_date else '',
                    'supplier': str(pos.supplier),
                })
            return JsonResponse({'data': data})
        else:
            return JsonResponse({'data': 'No Products found ...'})

    return JsonResponse({})







def get_active_users(request):
    active_users = models.Employee.objects.filter(is_active=True).values('first_name' , 'last_name')
    return JsonResponse({'active_users': list(active_users)})




from django.shortcuts import render, redirect, get_object_or_404
from .models import Product, ProductAttribute
from .forms import ProductForm, ProductAttributeFormSet

def product_create(request):
    """
    إنشاء منتج جديد مع خصائصه (Attributes) باستخدام formset
    """
    # جلب معرف الموظف من السيشن (إذا موجود)
    employee_id = request.session.get('employee_id')

    if request.method == "POST":
        form = ProductForm(request.POST)
        formset = ProductAttributeFormSet(request.POST, queryset=ProductAttribute.objects.none())
        
        if form.is_valid() and formset.is_valid():
            # حفظ المنتج وربطه بالموظف
            product = form.save(commit=False)
            if employee_id:
                product.employee_id = employee_id
            product.save()

            # حفظ الخصائص وربطها بالمنتج
            formset.instance = product
            formset.save()

            return redirect('product_list')  # تأكد أن لديك URL اسمه product_list
    else:
        form = ProductForm()
        formset = ProductAttributeFormSet(queryset=ProductAttribute.objects.none())

    return render(request, 'products/product_form.html', {
        'form': form,
        'formset': formset
    })


def product_list(request):
    """
    عرض قائمة المنتجات مع خصائصها
    """
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    all_products = Product.objects.all().count()
    search_query = request.GET.get('search', '').strip()
    products_qs = Product.objects.all().order_by('-id')
    if search_query:
        products_qs = products_qs.filter(product_name__icontains=search_query)
    paginator = Paginator(products_qs, 10)  # 10 products per page
    page = request.GET.get('page', 1)
    try:
        products = paginator.page(page)
    except PageNotAnInteger:
        products = paginator.page(1)
    except EmptyPage:
        products = paginator.page(paginator.num_pages)
    return render(request, 'products/product_list.html', {'products': products, 'search_query': search_query , 'all_products': all_products})




def product_update(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == "POST":
        form = ProductForm(request.POST, instance=product)
        formset = ProductAttributeFormSet(request.POST, instance=product)
        if form.is_valid() and formset.is_valid():
            form.save()
            formset.save()
            return redirect('product_list')
    else:
        form = ProductForm(instance=product)
        formset = ProductAttributeFormSet(instance=product)
    return render(request, 'products/product_form.html', {'form': form, 'formset': formset})


def product_delete(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == "POST":
        product.delete()
        return redirect('product_list')
    return render(request, 'products/product_confirm_delete.html', {'product': product})


def import_products_excel(request):
    """
    Import products from Excel file (.xlsx)
    Expected columns: product_name, quantity, purchasing_price, sale_price, category, expiry_date, isbn, production_date, author, supplier
    """
    def parse_date(value):
        """Parse date from various formats: string, datetime object, number (Excel serial), or None"""
        if not value:
            return None
        
        # If it's already a date object, return it
        if hasattr(value, 'date') and not isinstance(value, str):
            return value.date()
        
        # If it's a number, check if it's a year or Excel serial date
        if isinstance(value, (int, float)):
            # If it's a reasonable year (1900-2100), treat as year
            if 1900 <= value <= 2100:
                try:
                    from datetime import datetime
                    return datetime(int(value), 1, 1).date()
                except:
                    return None
            # Otherwise, treat as Excel serial date
            else:
                try:
                    from datetime import datetime, timedelta
                    # Excel serial date: 1 = 1900-01-01
                    base_date = datetime(1899, 12, 30)  # Excel's base date
                    return (base_date + timedelta(days=int(value))).date()
                except:
                    return None
        
        # If it's a string, try to parse it
        if isinstance(value, str):
            value = value.strip()
            if not value:
                return None
            
            # Try YYYY-MM-DD format first
            try:
                from datetime import datetime
                return datetime.strptime(value, '%Y-%m-%d').date()
            except ValueError:
                pass
            
            # Try other common formats (year parsing is handled above for numbers)
            
            # Try other common formats
            formats_to_try = [
                '%m/%d/%Y', '%d/%m/%Y', '%Y/%m/%d', '%d-%m-%Y', '%m-%d-%Y',
                '%Y-%m-%d', '%d/%m/%y', '%m/%d/%y', '%y-%m-%d', '%y/%d/%m',
                '%Y%m%d', '%d%m%Y', '%m%d%Y',
                '%b %d, %Y', '%d %b %Y', '%B %d, %Y', '%d %B %Y',  # Jan 15, 2024
                '%Y/%m/%d', '%d/%m/%Y', '%m/%d/%Y',  # with slashes
            ]
            
            for fmt in formats_to_try:
                try:
                    return datetime.strptime(value, fmt).date()
                except ValueError:
                    continue
            
            # If all fail, return None (will cause error later)
            return None
        
        # For any other type, try to convert to string and parse
        try:
            str_value = str(value).strip()
            if str_value:
                from datetime import datetime
                return datetime.strptime(str_value, '%Y-%m-%d').date()
        except:
            pass
        
        return None
    
    if request.method == 'POST':
        if 'excel_file' not in request.FILES:
            messages.error(request, _('Please select an Excel file to upload.'))
            return redirect('import_products_excel')
        
        excel_file = request.FILES['excel_file']
        if not excel_file.name.endswith('.xlsx'):
            messages.error(request, _('Please upload a valid Excel file (.xlsx).'))
            return redirect('import_products_excel')
        
        
        try:
            from openpyxl import load_workbook
            wb = load_workbook(excel_file)
            ws = wb.active
            
            # Skip header row
            rows = list(ws.iter_rows(values_only=True))[1:]
            
            imported_count = 0
            errors = []
            
            for row_num, row in enumerate(rows, start=2):  # Start from 2 because we skipped header
                try:
                    # Extract data from row
                    product_name = row[0] if len(row) > 0 and row[0] else None
                    quantity = row[1] if len(row) > 1 and row[1] is not None else 0
                    purchasing_price = row[2] if len(row) > 2 and row[2] is not None else None
                    sale_price = row[3] if len(row) > 3 and row[3] is not None else None
                    category = row[4] if len(row) > 4 and row[4] else None
                    # expiry_date_raw = row[5] if len(row) > 5 and row[5] else None
                    isbn = row[5] if len(row) > 5 and row[5] else None
                    publisher = row[6] if len(row) > 6 and row[6] else None
                    production_date_raw = row[7] if len(row) > 7 and row[7] else None
                    author = row[8] if len(row) > 8 and row[8] else None
                    supplier = row[9] if len(row) > 9 and row[9] else None
                    
                    # Validate required fields
                    if not product_name :
                        errors.append(_("Row %(row_num)d: Missing required fields (product_name)") % {'row_num': row_num})
                        continue
                    
                    # Parse dates
                    
                    # expiry_date = parse_date(expiry_date_raw)
                    production_date = parse_date(production_date_raw)
                    
                    # If date parsing failed but value was provided, add error
                    
                    # if expiry_date_raw and expiry_date is None:
                    #     errors.append(_("Row %(row_num)d: Invalid expiry_date format (use YYYY-MM-DD, YYYY, or other common formats)") % {'row_num': row_num})
                    #     continue
                    if production_date_raw and production_date is None:
                        errors.append(_("Row %(row_num)d: Invalid production_date format (use YYYY-MM-DD, YYYY, or other common formats)") % {'row_num': row_num})
                        continue
                    
                    # Convert prices to decimal
                    try:
                        purchasing_price = float(purchasing_price)
                    except (ValueError, TypeError):
                        errors.append(_("Row %(row_num)d: Invalid purchasing_price") % {'row_num': row_num})
                        continue
                    
                    if sale_price is not None:
                        try:
                            sale_price = float(sale_price)
                        except (ValueError, TypeError):
                            errors.append(_("Row %(row_num)d: Invalid sale_price") % {'row_num': row_num})
                            continue
                    
                    # Convert quantity to int
                    try:
                        quantity = int(quantity)
                    except (ValueError, TypeError):
                        quantity = 0
                    
                    # Check if ISBN already exists
                    if models.Product.objects.filter(isbn=isbn).exists():
                        errors.append(_("Row %(row_num)d: Product with ISBN '%(isbn)s' already exists") % {'row_num': row_num, 'isbn': isbn})
                        continue
                    
                    # Create product
                    product = models.Product.objects.create(
                        product_name=product_name,
                        quantity=quantity,
                        purchasing_price=purchasing_price,
                        sale_price=sale_price,
                        category=category,
                        # expiry_date=expiry_date,
                        isbn=isbn,
                        publisher=publisher,
                        production_date=production_date,
                        author=author,
                        supplier=supplier,
                        employee=request.session.get('employee_id') and models.Employee.objects.get(id=request.session['employee_id']) or None
                    )
                    
                    imported_count += 1
                    
                except Exception as e:
                    errors.append(_("Row %(row_num)d: %(error)s") % {'row_num': row_num, 'error': str(e)})
                    continue
            
            # Check for errors - reject entire file if any errors found
            if errors:
                messages.error(request, _('Import failed due to errors in the Excel file. Please fix the following issues and try again:'))
                for error in errors:
                    messages.error(request, error)
                return redirect('import_products_excel')
            
            # Show results only if no errors
            if imported_count > 0:
                messages.success(request, _('Successfully imported %(count)d products.') % {'count': imported_count})
            
            return redirect('product_list')
            
        except Exception as e:
            messages.error(request, _('Error processing Excel file: %(error)s') % {'error': str(e)})
            return redirect('import_products_excel')
    
    # GET request - show upload form
    return render(request, 'products/import_excel.html')


def download_sample_excel(request):
    """
    Generate and download a sample Excel file for product import
    """
    from openpyxl import Workbook
    from django.http import HttpResponse
    from io import BytesIO
    from openpyxl.styles import Font, PatternFill

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Products'

    # Add headers
    #  _('Expiry Date'), was removed from headers and sample data
    headers = [_('Product Name'), _('Quantity'), _('Purchasing Price'), _('Sale Price'), _('Category'), _('ISBN'), _('Publisher'), _('Production Date'), _('Author'), _('Supplier')]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = cell.alignment.copy(horizontal='center')
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    # Set column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 15
    sample_data = [
        ['Sample Book 1', 10, 15.50, 25.00, 'Fiction', '9781234567890', 'ABC Publishers', '2024-01-15', 'John Doe', 'ABC Publishers'],
        ['Sample Book 2', 5, 12.75, 20.00, 'Non-Fiction', '9780987654321', 'XYZ Books', '2024-01-15', 'Jane Smith', 'XYZ Books'],
        ['Sample Magazine', 20, 5.00, 8.50, 'Magazine', '9781122334455', 'News Corp', '2024-03-01', '', 'News Corp']
    ]

    for row_num, row_data in enumerate(sample_data, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)

    # Save to BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Create response
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="sample_products_import.xlsx"'

    return response

#-------------------- Excel reports downloads ----------------------------

# Empty products report
def download_empty_products_excel(request):
    """
    Generate and download Excel file for empty products report
    """
    from openpyxl import Workbook
    from django.http import HttpResponse
    from io import BytesIO
    from openpyxl.styles import Font, PatternFill

    # Get empty products
    products = models.Product.objects.filter(quantity=0)

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Empty Products'

    # Add headers
    headers = [_('Product Name'), _('Author'), _('Supplier'), _('Sale Price'),_('Publisher'), _('Production Date'), _('Quantity')]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = cell.alignment.copy(horizontal='center')
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    # Set column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 15
    for row_num, product in enumerate(products, 2):
        ws.cell(row=row_num, column=1, value=product.product_name)
        ws.cell(row=row_num, column=2, value=product.author)
        ws.cell(row=row_num, column=3, value=product.supplier)
        ws.cell(row=row_num, column=4, value=float(product.sale_price) if product.sale_price else 0)
        ws.cell(row=row_num, column=5, value=product.publisher)
        ws.cell(row=row_num, column=6, value=product.production_date.strftime('%Y') if product.production_date else '')
        ws.cell(row=row_num, column=7, value=product.quantity)

    # Save to BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Create response
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="empty_products_report.xlsx"'

    return response

# Available products excel report
def download_stock_products_excel(request):
    """
    Generate and download Excel file for stock products report
    """
    from openpyxl import Workbook
    from django.http import HttpResponse
    from io import BytesIO
    from openpyxl.styles import Font, PatternFill

    # Get stock products
    products = models.Product.objects.filter(quantity__gt=0).order_by('-updated_at')

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Stock Products'

    # Add headers
    headers = [_('Product Name'), _('Author'), _('Supplier'), _('Sale Price'),_('Publisher'), _('Production Date'), _('Quantity')]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = cell.alignment.copy(horizontal='center')
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid", )

    # Set column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 15
    for row_num, product in enumerate(products, 2):
        ws.cell(row=row_num, column=1, value=product.product_name)
        ws.cell(row=row_num, column=2, value=product.author)
        ws.cell(row=row_num, column=3, value=product.supplier)
        ws.cell(row=row_num, column=4, value=float(product.sale_price) if product.sale_price else 0).font = Font(bold=True)
        ws.cell(row=row_num, column=5, value=product.publisher)
        ws.cell(row=row_num, column=6, value=product.production_date.strftime('%Y') if product.production_date else '')
        ws.cell(row=row_num, column=7, value=product.quantity)

    # Save to BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Create response
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="stock_products_report.xlsx"'

    return response

###########################################################################

from django.utils.translation import gettext as gettext
from django.db.models import Sum
from django.db import transaction
from datetime import date
from decimal import Decimal
from openpyxl import load_workbook
from django.contrib import messages
from django.shortcuts import render, redirect
from . import models


def import_purchase_invoices_excel(request):
    """
    Import purchase invoices from Excel file (.xlsx) using submet_purchase_order logic.
    Generates **one invoice per supplier + employee + payment_method + currency**.
    """
    if request.method == 'POST':
        if 'excel_file' not in request.FILES:
            messages.error(request, gettext('Please select an Excel file to upload.'))
            return redirect('import_purchase_invoices_excel')
        
        excel_file = request.FILES['excel_file']
        if not excel_file.name.endswith('.xlsx'):
            messages.error(request, gettext('Please upload a valid Excel file (.xlsx).'))
            return redirect('import_purchase_invoices_excel')
        
        try:
            wb = load_workbook(excel_file)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))[1:]  # Skip header

            errors = []
            imported_invoices = 0
            imported_items = 0

            # Get base currency
            company = models.CompanyProfile.objects.first()
            base_currency_id = company.base_currency_id if company else None

            # Group rows by supplier + employee + payment_method + currency
            invoice_groups = {}

            for row_num, row in enumerate(rows, start=2):
                try:
                    product_name = row[0] if len(row) > 0 else None
                    quantity = row[1] if len(row) > 1 else None
                    unit_price = row[2] if len(row) > 2 else None
                    product_isbn = row[3] if len(row) > 3 else None
                    supplier_name = row[4] if len(row) > 4 else None
                    payment_method = row[5] if len(row) > 5 else 'cash'
                    currency_code = row[6] if len(row) > 6 else None

                    # Validate required fields
                    if not all([product_name, quantity, unit_price, product_isbn, supplier_name, currency_code]):
                        errors.append(gettext("Row %(row)d: Missing required fields") % {'row': row_num})
                        continue

                    # Convert quantity and unit_price
                    try:
                        quantity = int(quantity)
                        unit_price = float(unit_price)
                        total_price = quantity * unit_price
                    except Exception:
                        errors.append(gettext("Row %(row)d: Invalid quantity or unit_price") % {'row': row_num})
                        continue

                    # Validate product
                    try:
                        product = models.Product.objects.get(isbn=product_isbn)
                        if product.product_name != product_name:
                            errors.append(gettext("Row %(row)d: Product name '%(pname)s' does not match '%(expected)s'") %
                                          {'row': row_num, 'pname': product_name, 'expected': product.product_name})
                            continue
                    except models.Product.DoesNotExist:
                        errors.append(gettext("Row %(row)d: Product with ISBN '%(isbn)s' not found") %
                                      {'row': row_num, 'isbn': product_isbn})
                        continue

                    # Validate currency
                    try:
                        currency = models.Currency.objects.get(code=currency_code)
                    except models.Currency.DoesNotExist:
                        errors.append(gettext("Row %(row)d: Currency '%(currency)s' not found") %
                                      {'row': row_num, 'currency': currency_code})
                        continue

                    # Get employee
                    employee_id = request.session.get('employee_id')
                    try:
                        employee = models.Employee.objects.get(id=employee_id)
                    except models.Employee.DoesNotExist:
                        errors.append(gettext("Row %(row)d: Employee with ID '%(id)s' not found") %
                                      {'row': row_num, 'id': employee_id})
                        continue

                    # Find or create supplier
                    supplier, _ = models.Supplier.objects.get_or_create(
                        name=supplier_name, defaults={'phone': '', 'email': '', 'contact_info': ''}
                    )

                    # Create invoice key
                    invoice_key = f"{supplier.id}_{employee.id}_{payment_method}_{currency.id}"

                    if invoice_key not in invoice_groups:
                        invoice_groups[invoice_key] = {
                            'supplier': supplier,
                            'employee': employee,
                            'payment_method': payment_method,
                            'currency': currency,
                            'items': []
                        }

                    invoice_groups[invoice_key]['items'].append({
                        'product': product,
                        'quantity': quantity,
                        'unit_price': unit_price,
                        'total_price': total_price
                    })

                except Exception as e:
                    errors.append(gettext("Row %(row)d: %(error)s") % {'row': row_num, 'error': str(e)})
                    continue

            if errors:
                messages.error(request, gettext('Import failed due to errors in the Excel file. Please fix and try again:'))
                for err in errors:
                    messages.error(request, err)
                return redirect('import_purchase_invoices_excel')

            # ================== CREATE PURCHASE INVOICES ==================
            for key, data in invoice_groups.items():
                try:
                    with transaction.atomic():
                        employee = data['employee']
                        supplier = data['supplier']
                        payment_method = data['payment_method']
                        currency = data['currency']
                        items = data['items']

                        # Create purchase order
                        purchase = models.create_purchase_order(employee.id, supplier_id=supplier.id)
                        purchase.invoice_pay_method = payment_method
                        purchase.currency_id = currency.id

                        # Exchange rate
                        exchange_rate_obj = models.ExchangeRate.objects.filter(
                            from_currency_id=currency.id,
                            to_currency_id=base_currency_id,
                            date=date.today()
                        ).first()
                        if exchange_rate_obj:
                            purchase.exchange_rate_to_base = exchange_rate_obj.rate
                        elif currency.id != base_currency_id:
                            purchase.exchange_rate_to_base = 1.0
                        else:
                            purchase.exchange_rate_to_base = 1.0
                        purchase.save()

                        total_amount = 0

                        # Add items and update products
                        from decimal import Decimal
                        for item in items:
                            product = item['product']
                            quantity = item['quantity']
                            unit_price = item['unit_price']
                            total_price = item['total_price']
                            

                            models.add_item_to_purchase_invoice(product.id, quantity, unit_price, total_price)
                            
                            # Update product quantity
                            product.quantity += quantity

                            # Update product purchasing price if currency != base
                            if currency.id != base_currency_id:
                                product.purchasing_price = Decimal(unit_price) * Decimal(purchase.exchange_rate_to_base)
                                messages.warning(request, gettext("Updated purchasing price for product '%(prod)s' to %(price).2f in base currency.") % {'prod': product.product_name, 'price': product.purchasing_price})
                            else:
                                product.purchasing_price = unit_price
                            product.save()

                            total_amount += total_price
                            imported_items += 1

                        # Update purchase totals
                        purchase.total_amount = Decimal(total_amount)
                        purchase.grand_total = Decimal(total_amount)
                        purchase.total_price_base = Decimal(total_amount) * Decimal(purchase.exchange_rate_to_base)
                        purchase.save()

                        # Update purchase_items with total_price_base
                        for pi in models.Purchase_item.objects.filter(purchase_id=purchase):
                            pi.currency_id = currency.id
                            pi.exchange_rate_to_base = Decimal(purchase.exchange_rate_to_base)
                            pi.total_price_base = Decimal(pi.total_price) * Decimal(purchase.exchange_rate_to_base)
                            pi.save()

                        imported_invoices += 1

                except Exception as e:
                    messages.warning(request, gettext("Error creating invoice for %(key)s: %(error)s") %
                                     {'key': key, 'error': str(e)})
                    continue

            if imported_invoices > 0:
                messages.success(request, gettext(
                    'Successfully imported %(invoices)d purchase invoices with %(items)d items.'
                ) % {'invoices': imported_invoices, 'items': imported_items})

            return redirect('display_purchases')

        except Exception as e:
            messages.error(request, gettext('Error processing Excel file: %(error)s') % {'error': str(e)})
            return redirect('import_purchase_invoices_excel')

    return render(request, 'purchases/import_excel.html')


################################################################################


def download_sample_purchase_excel(request):
    """
    Generate and download a sample Excel file for purchase invoice import
    """
    from openpyxl import Workbook
    from django.http import HttpResponse
    from io import BytesIO
    from openpyxl.styles import Font, PatternFill

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Purchase Invoices'

    # Add headers
    headers = [ _('Product Name'),_('Quantity'),_('Unit Price'),_('Product ISBN'),_('Supplier Name'), _('Payment Method'),_('Currency')]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = cell.alignment.copy(horizontal='center')
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    # Set column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 15
    sample_data = [
        ['Sample Book 1',10,15.50,'9781234567890','ABC Suppliers', 'cash','ILS' ],
        ['Sample Book 2',5,12.75,'9780987654321','ABC Suppliers',  'cash','ILS'],
        ['Sample Magazine',20,8.00,'9781122334455','XYZ Distributors',  'debts','USD'],
    ]

    for row_num, row_data in enumerate(sample_data, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)

    # Save to BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Create response
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="sample_purchase_invoices_import.xlsx"'

    return response


def sales_products_report(request):

    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    report_data = []
    total_sold_quantity = 0
    total_return_sale_quantity = 0
    total_remain_sold_quantity = 0
    grand_total_money_should_received = 0
    error_message = None

    # Only generate report if both dates are provided
    if from_date and to_date:
        sales = models.Sale_item.objects.filter(created_at__date__gte=from_date, created_at__date__lte=to_date)
        for sale in sales:
            product = sale.product
            product_name = product.product_name
            author_name = getattr(product, 'author', '')
            supplier_name = getattr(product, 'supplier', '')
            sale_price = sale.unit_price
            production_date = getattr(product, 'production_date', '')
            sold_quantity = sale.quantity
            # Get return sale quantity for this sale item
            return_sale_quantity = models.SaleReturnItem.objects.filter(original_item=sale).aggregate(total=Sum('quantity'))['total'] or 0
            remain_sold_quantity = sold_quantity - return_sale_quantity
            money_should_received = remain_sold_quantity * sale_price

            report_data.append({
                'product_name': product_name,
                'author_name': author_name,
                'supplier_name': supplier_name,
                'sale_price': sale_price,
                'production_date': production_date,
                'sold_quantity': sold_quantity,
                'return_sale_quantity': return_sale_quantity,
                'remain_sold_quantity': remain_sold_quantity,
                'money_should_received': money_should_received,
            })

            total_sold_quantity += sold_quantity
            total_return_sale_quantity += return_sale_quantity
            total_remain_sold_quantity += remain_sold_quantity
            grand_total_money_should_received += money_should_received

        # Pagination
        from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
        page = request.GET.get('page', 1)
        paginator = Paginator(report_data, 20)  # 20 items per page
        try:
            report_page = paginator.page(page)
        except PageNotAnInteger:
            report_page = paginator.page(1)
        except EmptyPage:
            report_page = paginator.page(paginator.num_pages)
    elif from_date or to_date:
        error_message = "Please select both From Date and To Date."
        report_page = None
    else:
        report_page = None

    context = {
        'report_data': report_page,
        'from_date': from_date,
        'to_date': to_date,
        'total_sold_quantity': total_sold_quantity,
        'total_return_sale_quantity': total_return_sale_quantity,
        'total_remain_sold_quantity': total_remain_sold_quantity,
        'grand_total_money_should_received': grand_total_money_should_received,
        'error_message': error_message,
    }
    return render(request, 'sales_products_report.html', context)


def download_sales_products_report_excel(request):
    """
    Generate and download Excel file for sales products report
    """
    from openpyxl import Workbook
    from django.http import HttpResponse
    from io import BytesIO
    from openpyxl.styles import Font, PatternFill

    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')

    if not from_date or not to_date:
        from django.contrib import messages
        messages.error(request, "Please provide both From Date and To Date.")
        return redirect('sales_products_report')

    # Generate report data (same logic as sales_products_report)
    report_data = []
    sales = models.Sale_item.objects.filter(created_at__date__gte=from_date, created_at__date__lte=to_date)
    for sale in sales:
        product = sale.product
        product_name = product.product_name
        author_name = getattr(product, 'author', '')
        supplier_name = getattr(product, 'supplier', '')
        sale_price = sale.unit_price
        production_date = getattr(product, 'production_date', '')
        sold_quantity = sale.quantity
        # Get return sale quantity for this sale item
        return_sale_quantity = models.SaleReturnItem.objects.filter(original_item=sale).aggregate(total=Sum('quantity'))['total'] or 0
        remain_sold_quantity = sold_quantity - return_sale_quantity
        money_should_received = remain_sold_quantity * sale_price

        report_data.append({
            'product_name': product_name,
            'author_name': author_name,
            'supplier_name': supplier_name,
            'sale_price': sale_price,
            'production_date': production_date,
            'sold_quantity': sold_quantity,
            'return_sale_quantity': return_sale_quantity,
            'remain_sold_quantity': remain_sold_quantity,
            'money_should_received': money_should_received,
        })

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Sales Products Report'

    # Add headers
    headers = [
        _('Product Name'), _('Author'), _('Supplier'), _('Sale Price'),
        _('Production Date'), _('Sold Quantity'), _('Return Quantity'),
        _('Remaining Quantity'), _('Money Should Received')
    ]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = cell.alignment.copy(horizontal='center')
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    # Set column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 15

    # Add data
    for row_num, item in enumerate(report_data, 2):
        ws.cell(row=row_num, column=1, value=item['product_name'])
        ws.cell(row=row_num, column=2, value=item['author_name'])
        ws.cell(row=row_num, column=3, value=item['supplier_name'])
        ws.cell(row=row_num, column=4, value=float(item['sale_price']) if item['sale_price'] else 0)
        ws.cell(row=row_num, column=5, value=str(item['production_date'].strftime('%Y')) if item['production_date'] else '')
        ws.cell(row=row_num, column=6, value=item['sold_quantity'])
        ws.cell(row=row_num, column=7, value=item['return_sale_quantity'])
        ws.cell(row=row_num, column=8, value=item['remain_sold_quantity'])
        ws.cell(row=row_num, column=9, value=float(item['money_should_received']) if item['money_should_received'] else 0)

    # Add grand total row
    grand_total_row = len(report_data) + 2
    ws.cell(row=grand_total_row, column=8, value=_('Grand Total')).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws.cell(row=grand_total_row, column=8).font = Font(bold=True)
    ws.cell(row=grand_total_row, column=9, value=sum(float(item['money_should_received']) if item['money_should_received'] else 0 for item in report_data)).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws.cell(row=grand_total_row, column=9).font = Font(bold=True)

    # Save to BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Create response
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="sales_products_report_{from_date}_to_{to_date}.xlsx"'

    return response







def _get_Stock_Out_cart(request):
    return request.session.get('stock_out_cart', [])


def _save_Stock_Out_cart(request, cart):
    request.session['stock_out_cart'] = cart
    request.session.modified = True


def delete_product_from_stock_out(request):
    """
    AJAX endpoint: delete a product from purchase cart by product_id
    Returns JSON with updated grand_total and cart items
    """
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid method'}, status=405)
    
    product_id = request.POST.get('product_id')
    if not product_id:
        return JsonResponse({'status': 'error', 'message': 'No product_id provided'}, status=400)
    
    try:
        product_id = int(product_id)
    except (ValueError, TypeError):
        return JsonResponse({'status': 'error', 'message': 'Invalid product_id'}, status=400)
    
    cart = _get_Stock_Out_cart(request)
    # Find and remove the item by product_id
    original_length = len(cart)
    cart = [item for item in cart if int(item.get('product_id')) != product_id]
    
    if len(cart) == original_length:
        # Item not found
        return JsonResponse({'status': 'error', 'message': 'Product not found in cart'}, status=404)
    
    _save_Stock_Out_cart(request, cart)
    
    # Recalculate grand total
    grand_total = sum(float(item.get('total_price', 0)) for item in cart)
    
    return JsonResponse({'status': 'ok', 'grand_total': grand_total, 'items': cart})

def update_stock_out_cart(request):
    """
    Persist edited stock-out cart rows to session.
    Accepts JSON payload: { items: [ {product_id, quantity, purchase_price} ] }
    Returns updated items and grand_total.
    """
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid method'}, status=405)

    items_payload = None
    if request.content_type == 'application/json':
        import json
        try:
            body = json.loads(request.body or '{}')
            items_payload = body.get('items')
        except Exception:
            return JsonResponse({'status': 'error', 'message': _('Invalid payload')}, status=400)
    else:
        # Fallback: accept form-encoded 'items' JSON string
        items_str = request.POST.get('items')
        if items_str:
            import json
            try:
                items_payload = json.loads(items_str)
            except Exception:
                items_payload = None

    if not isinstance(items_payload, list):
        return JsonResponse({'status': 'error', 'message': _('No items provided')}, status=400)

    new_cart = []
    for it in items_payload:
        product_id = it.get('product_id')
        quantity = it.get('quantity')
        purchase_price = it.get('purchase_price')
        try:
            product_id = int(product_id)
        except (ValueError, TypeError):
            return JsonResponse({'status': 'error', 'message': _('Invalid product_id')}, status=400)
        try:
            quantity = int(quantity)
        except (ValueError, TypeError):
            quantity = 0
        try:
            purchase_price = float(purchase_price)
        except (ValueError, TypeError):
            purchase_price = 0.0

        if quantity <= 0:
            return JsonResponse({'status': 'error', 'message': _('Quantity must be positive')}, status=400)

        try:
            prod = models.Product.objects.get(id=product_id)
        except models.Product.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': _('Product not found')}, status=404)

        # Stock validations consistent with add-to-cart
        try:
            available_qty = int(prod.quantity)
        except Exception:
            available_qty = prod.quantity or 0
        if available_qty < 0:
            msg = _('Product "%(prod)s" is out of stock.') % {'prod': prod.product_name}
            return JsonResponse({'status': 'error', 'message': msg}, status=400)
        if available_qty < quantity:
            msg = _('Not enough stock for product "%(prod)s". Available: %(avail)d') % {'prod': prod.product_name, 'avail': available_qty}
            return JsonResponse({'status': 'error', 'message': msg}, status=400)

        total_price = quantity * purchase_price
        new_cart.append({
            'product_name': prod.product_name,
            'product_id': product_id,
            'quantity': quantity,
            'purchase_price': purchase_price,
            'total_price': total_price,
        })

    _save_Stock_Out_cart(request, new_cart)

    grand_total = sum(float(item.get('total_price', 0)) for item in new_cart)
    return JsonResponse({'status': 'ok', 'grand_total': grand_total, 'items': new_cart})

############################################
def display_Stock_Out_voucher(request):
    # calculate grand total (session-backed)
    Stock_Out_cart = _get_Stock_Out_cart(request)
    grand_total = sum(float(item.get('total_price', 0)) for item in Stock_Out_cart)
    # paginate recent purchase invoices
    stock_out_vouchers_qs = models.get_all_Stock_Out_Vouchers()
    stock_out_vouchers_page_number = request.GET.get('vouchers_page')
    stock_out_vouchers_paginator = Paginator(stock_out_vouchers_qs, 10)  # 10 invoices per page
    stock_out_vouchers_page = stock_out_vouchers_paginator.get_page(stock_out_vouchers_page_number)
    

    context = {
            # alias to match existing template keys
            'purchases_order': Stock_Out_cart,
            'stock_outs_order': Stock_Out_cart,
            'products': models.get_all_products(),
            # recent vouchers list + aliases used by template
            'stock_out_vouchers': stock_out_vouchers_page,
            'invoices': stock_out_vouchers_page,
            'employee': models.get_employee_by_id(request.session['employee_id']),
            'grand_total': grand_total,
            'stock_out_vouchers_paginator': stock_out_vouchers_paginator,
            'invoices_paginator': stock_out_vouchers_paginator,
            #'suppliers': models.Supplier.objects.all(),
            'customers': models.Customer.objects.all(),
            'currencies': models.Currency.objects.all(),
            'voucher' : stock_out_vouchers_qs,
        }
    return render(request , 'stock_Out_Voucher.html' ,context)
############################################

# ------------------- Stock Out Voucher Cart Ops -------------------
from decimal import Decimal

def add_product_to_stock_out(request):
    """
    Add product to session-backed stock out cart. Supports normal POST and AJAX/JSON.
    Returns JSON when called via AJAX, else redirects back to stock-out page.
    """
    data = None
    if request.content_type == 'application/json':
        import json
        try:
            data = json.loads(request.body)
        except Exception:
            return JsonResponse({'status': 'error', 'message': _('Invalid payload')}, status=400)

    product_name = (data.get('product_name') if data else request.POST.get('product_name'))
    quantity = int((data.get('quantity') if data else request.POST.get('quantity', 0)) or 0)

    if not product_name or quantity <= 0:
        msg = _('Product name and positive quantity are required.')
        if request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.content_type == 'application/json':
            return JsonResponse({'status': 'error', 'message': msg}, status=400)
        messages.error(request, msg)
        return redirect('/stock_out_voucher')

    try:
        prod = models.Product.objects.get(product_name=product_name)
    except models.Product.DoesNotExist:
        msg = _('Product not found')
        if request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.content_type == 'application/json':
            return JsonResponse({'status': 'not_found', 'message': msg})
        messages.error(request, msg)
        return redirect('/stock_out_voucher')

    # Stock validations (consider existing quantity in cart for this product)
    try:
        available_qty = int(prod.quantity)
    except Exception:
        available_qty = prod.quantity or 0
    # existing quantity of this product already in cart
    existing_qty_in_cart = 0
    try:
        existing_qty_in_cart = sum(int(i.get('quantity', 0)) for i in _get_Stock_Out_cart(request) if int(i.get('product_id', 0)) == int(prod.id))
    except Exception:
        existing_qty_in_cart = 0
    if available_qty < 0:
        msg = _('Product "%(prod)s" is out of stock.') % {'prod': product_name}
        if request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.content_type == 'application/json':
            return JsonResponse({'status': 'error', 'message': msg}, status=400)
        messages.error(request, msg)
        return redirect('/stock_out_voucher')
    # validate against total desired quantity (existing + new)
    total_desired = existing_qty_in_cart + quantity
    if available_qty < total_desired:
        msg = _('Not enough stock for product "%(prod)s". Available: %(avail)d') % {'prod': product_name, 'avail': available_qty}
        if request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.content_type == 'application/json':
            return JsonResponse({'status': 'error', 'message': msg}, status=400)
        messages.error(request, msg)
        return redirect('/stock_out_voucher')

    product_id = prod.id
    unit_price = float(prod.purchasing_price) if prod.purchasing_price is not None else 0.0
    total_price = quantity * unit_price
    cart = _get_Stock_Out_cart(request)
    for item in cart:
        if int(item.get('product_id')) == int(product_id):
            item['quantity'] = int(item.get('quantity', 0)) + quantity
            try:
                item['total_price'] = int(item['quantity']) * float(item.get('purchase_price', unit_price))
            except Exception:
                item['total_price'] = 0
            break
    else:
        cart.append({
            'product_name': product_name,
            'product_id': product_id,
            'quantity': quantity,
            # keep same key name used by template/JS
            'purchase_price': unit_price,
            'total_price': total_price,
            
        })
    _save_Stock_Out_cart(request, cart)

    grand_total = sum(float(item.get('total_price', 0)) for item in cart)
    if request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.content_type == 'application/json':
        return JsonResponse({'status': 'ok', 'grand_total': grand_total, 'items': cart})
    return redirect('/stock_out_voucher')


def submit_stock_out_order(request):
    """
    Persist current stock-out cart to DB as Stock_Out_Voucher and Stock_Out_Voucher_item rows
    without affecting Product.quantity.
    """
    cart = _get_Stock_Out_cart(request)
    if not cart:
        messages.error(request, _('Please add at least one product to stock-out voucher!'))
        return redirect('/stock_out_voucher')

    currency_id = request.POST.get('currency_id')
    customer_id = request.POST.get('customer_id')

    if not customer_id:
        messages.error(request, _('Please select a customer for the stock-out voucher!'))
        return redirect('/stock_out_voucher')

    pay_method = request.POST.get('invoice_pay_method', 'cash')

    # Create voucher
    employee_id = request.session.get('employee_id')
    if not employee_id:
        messages.error(request, _('You must be signed in as an employee.'))
        return redirect('/')

    voucher = models.Stock_Out_Voucher.objects.create(
        employee_id=employee_id,
        customer_id=customer_id or None,
        invoice_pay_method=pay_method,
    )

    # Set currency and exchange rate (use base -> selected currency logic)
    if currency_id:
        voucher.currency_id = currency_id
        try:
            exchange_rate_obj = models.ExchangeRate.objects.filter(
                from_currency_id=models.CompanyProfile.objects.first().base_currency_id if models.CompanyProfile.objects.first() and models.CompanyProfile.objects.first().base_currency else None,
                to_currency__id=currency_id,
                date=date.today()
            ).first()
            if exchange_rate_obj:
                voucher.exchange_rate_to_base = exchange_rate_obj.rate
            elif currency_id != (models.CompanyProfile.objects.first().base_currency_id if models.CompanyProfile.objects.first() else None):
                voucher.exchange_rate_to_base = 1.0
        except Exception as e:
            messages.warning(request, _("Could not set exchange rate: %(err)s") % {'err': str(e)})
    voucher.save()

    # Create voucher items and compute totals
    total_sum = Decimal('0.0')
    total_sum_base = Decimal('0.0')
    for key in cart:
        product_id = key.get('product_id')
        quantity = int(key.get('quantity'))
        unit_price = Decimal(str(key.get('purchase_price') or 0))
        total_price = Decimal(str(key.get('total_price') or 0))

        item = models.Stock_Out_Voucher_item.objects.create(
            Stock_Out_Voucher_id=voucher,
            product_id=product_id,
            quantity=quantity,
            unit_price=unit_price,
            total_price=total_price,
            currency_id=currency_id or None,
            exchange_rate_to_base=voucher.exchange_rate_to_base or Decimal('1.0'),
            total_price_base=(total_price * Decimal(str(voucher.exchange_rate_to_base or 1)))
        )
        total_sum += total_price
        total_sum_base += item.total_price_base

    voucher.grand_total = total_sum
    voucher.total_amount = total_sum
    voucher.total_price_base = total_sum_base
    voucher.save()

    # Clear cart and notify
    _save_Stock_Out_cart(request, [])
    messages.success(request, _('Stock-Out Voucher Created Successfully!'), extra_tags='add_invoice')
    return redirect('/stock_out_voucher')


def view_stock_out_invoice(request, id):
    try:
        voucher = models.Stock_Out_Voucher.objects.get(id=id)
    except models.Stock_Out_Voucher.DoesNotExist:
        messages.error(request, _('Voucher not found.'))
        return redirect('/stock_out_voucher')
    items = voucher.Stock_Out_Voucher_items.select_related('product').all()
    context = {
        'voucher': voucher,
        'items': items,
        'total': voucher.grand_total,
        'company': models.get_company_profile(),
    }
    return render(request, 'stock_out_voucher_detail.html', context)

######### applying the edit for stock out invoice edit view (in cart_session) ############
def stock_out_invoice_edit(request, id):
    try:
        voucher = models.Stock_Out_Voucher.objects.get(id=id)
    except models.Stock_Out_Voucher.DoesNotExist:
        messages.error(request, _('Voucher not found.'))
        return redirect('/stock_out_voucher')
    items = voucher.Stock_Out_Voucher_items.select_related('product').all()
    context = {
        'voucher': voucher,
        'items': items,
        'products': models.Product.objects.all(),
    }
    return render(request, 'stock_out_voucher_edit.html', context)
######### applying the edit for stock out invoice edit view (in cart_session) ############



def stock_out_invoice_add_item(request, id):
    if request.method != 'POST':
        return redirect(f'/stock_out_invoice/{id}/edit')
    try:
        voucher = models.Stock_Out_Voucher.objects.get(id=id)
    except models.Stock_Out_Voucher.DoesNotExist:
        messages.error(request, _('Voucher not found.'))
        return redirect('/stock_out_voucher')

    product_id = request.POST.get('product_id')
    quantity = int(request.POST.get('quantity', 0) or 0)
    unit_price = request.POST.get('unit_price')
    if not product_id or quantity <= 0:
        messages.error(request, _('Product and positive quantity are required.'))
        return redirect(f'/stock_out_invoice/{id}/edit')
    try:
        prod = models.Product.objects.get(id=product_id)
    except models.Product.DoesNotExist:
        messages.error(request, _('Product not found.'))
        return redirect(f'/stock_out_invoice/{id}/edit')
    # Stock validation: check available against existing voucher quantity + new qty
    try:
        available_qty = int(prod.quantity)
    except Exception:
        available_qty = prod.quantity or 0
    if available_qty < 0:
        messages.error(request, _('Product "%(prod)s" is out of stock.') % {'prod': prod.product_name})
        return redirect(f'/stock_out_invoice/{id}/edit')
    try:
        existing_voucher_qty = voucher.Stock_Out_Voucher_items.filter(product_id=product_id).aggregate(q=Sum('quantity')).get('q') or 0
        existing_voucher_qty = int(existing_voucher_qty)
    except Exception:
        existing_voucher_qty = 0
    total_desired = existing_voucher_qty + quantity
    if available_qty < total_desired:
        messages.error(request, _('Not enough stock for product "%(prod)s". Available: %(avail)d') % {'prod': prod.product_name, 'avail': available_qty})
        return redirect(f'/stock_out_invoice/{id}/edit')
    unit_price_val = Decimal(str(unit_price)) if unit_price is not None and unit_price != '' else Decimal(str(prod.purchasing_price or 0))
    # Merge with existing item instead of creating a new row
    existing_item = voucher.Stock_Out_Voucher_items.filter(product_id=prod.id).order_by('id').first()
    if existing_item:
        new_qty = int(existing_item.quantity or 0) + int(quantity)
        existing_item.quantity = new_qty
        existing_item.unit_price = unit_price_val
        existing_item.total_price = unit_price_val * Decimal(new_qty)
        existing_item.currency = voucher.currency
        existing_item.exchange_rate_to_base = voucher.exchange_rate_to_base
        existing_item.total_price_base = existing_item.total_price * Decimal(str(voucher.exchange_rate_to_base or 1))
        existing_item.save()
        merged_message = _('Item quantity updated.')
    else:
        total_price = unit_price_val * Decimal(quantity)
        models.Stock_Out_Voucher_item.objects.create(
            Stock_Out_Voucher_id=voucher,
            product=prod,
            quantity=quantity,
            unit_price=unit_price_val,
            total_price=total_price,
            currency=voucher.currency,
            exchange_rate_to_base=voucher.exchange_rate_to_base,
            total_price_base=(total_price * Decimal(str(voucher.exchange_rate_to_base or 1)))
        )
        merged_message = _('Item added.')
    # Recompute totals
    totals = voucher.Stock_Out_Voucher_items.aggregate(
        total=Sum('total_price'), base=Sum('total_price_base')
    )
    voucher.grand_total = totals.get('total') or 0
    voucher.total_amount = voucher.grand_total
    voucher.total_price_base = totals.get('base') or 0
    voucher.save()
    messages.success(request, merged_message)
    return redirect(f'/stock_out_invoice/{id}/edit')


def stock_out_invoice_delete_item(request, id):
    if request.method != 'POST':
        return redirect(f'/stock_out_invoice/{id}/edit')
    item_id = request.POST.get('item_id')
    try:
        item = models.Stock_Out_Voucher_item.objects.get(id=item_id, Stock_Out_Voucher_id_id=id)
    except models.Stock_Out_Voucher_item.DoesNotExist:
        messages.error(request, _('Item not found.'))
        return redirect(f'/stock_out_invoice/{id}/edit')
    voucher = item.Stock_Out_Voucher_id
    item.delete()
    totals = voucher.Stock_Out_Voucher_items.aggregate(
        total=Sum('total_price'), base=Sum('total_price_base')
    )
    voucher.grand_total = totals.get('total') or 0
    voucher.total_amount = voucher.grand_total
    voucher.total_price_base = totals.get('base') or 0
    voucher.save()
    messages.success(request, _('Item deleted.'))
    return redirect(f'/stock_out_invoice/{id}/edit')


def stock_out_invoice_update_items(request, id):
    """
    Persist inline edits on the stock-out voucher edit page.
    Accepts JSON: { items: [ { item_id, quantity, unit_price } ] }
    Applies validations against current product stock and updates totals.
    """
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid method'}, status=405)
    try:
        voucher = models.Stock_Out_Voucher.objects.get(id=id)
    except models.Stock_Out_Voucher.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': _('Voucher not found')}, status=404)

    # Parse payload
    items_payload = None
    if request.content_type == 'application/json':
        import json
        try:
            body = json.loads(request.body or '{}')
            items_payload = body.get('items')
        except Exception:
            return JsonResponse({'status': 'error', 'message': _('Invalid payload')}, status=400)
    else:
        items_str = request.POST.get('items')
        if items_str:
            import json
            try:
                items_payload = json.loads(items_str)
            except Exception:
                items_payload = None

    if not isinstance(items_payload, list):
        return JsonResponse({'status': 'error', 'message': _('No items provided')}, status=400)

    from decimal import Decimal as D
    # Apply edits
    for it in items_payload:
        item_id = it.get('item_id')
        quantity = it.get('quantity')
        unit_price = it.get('unit_price')
        try:
            item_id = int(item_id)
        except (ValueError, TypeError):
            return JsonResponse({'status': 'error', 'message': _('Invalid item_id')}, status=400)
        try:
            quantity = int(quantity)
        except (ValueError, TypeError):
            quantity = 0
        try:
            unit_price = D(str(unit_price))
        except Exception:
            unit_price = D('0')

        if quantity <= 0:
            return JsonResponse({'status': 'error', 'message': _('Quantity must be positive')}, status=400)

        try:
            item = models.Stock_Out_Voucher_item.objects.get(id=item_id, Stock_Out_Voucher_id=voucher)
        except models.Stock_Out_Voucher_item.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': _('Item not found')}, status=404)

        prod = item.product
        try:
            available_qty = int(prod.quantity)
        except Exception:
            available_qty = prod.quantity or 0
        if available_qty < 0:
            msg = _('Product "%(prod)s" is out of stock.') % {'prod': prod.product_name}
            return JsonResponse({'status': 'error', 'message': msg}, status=400)
        if available_qty < quantity:
            msg = _('Not enough stock for product "%(prod)s". Available: %(avail)d') % {'prod': prod.product_name, 'avail': available_qty}
            return JsonResponse({'status': 'error', 'message': msg}, status=400)

        # Update item
        item.quantity = quantity
        item.unit_price = unit_price
        item.total_price = unit_price * D(str(quantity))
        item.currency = voucher.currency
        item.exchange_rate_to_base = voucher.exchange_rate_to_base
        item.total_price_base = item.total_price * D(str(voucher.exchange_rate_to_base or 1))
        item.save()

    # Recompute voucher totals
    totals = voucher.Stock_Out_Voucher_items.aggregate(total=Sum('total_price'), base=Sum('total_price_base'))
    voucher.grand_total = totals.get('total') or 0
    voucher.total_amount = voucher.grand_total
    voucher.total_price_base = totals.get('base') or 0
    voucher.save()

    # Build response items
    resp_items = []
    for it in voucher.Stock_Out_Voucher_items.select_related('product').all():
        resp_items.append({
            'id': it.id,
            'product_id': it.product_id,
            'product_name': it.product.product_name if it.product_id else '',
            'quantity': it.quantity,
            'unit_price': float(it.unit_price) if it.unit_price is not None else 0,
            'total_price': float(it.total_price) if it.total_price is not None else 0,
        })

    return JsonResponse({'status': 'ok', 'grand_total': float(voucher.grand_total or 0), 'items': resp_items})
def convert_stock_out_to_sale(request, id):
    """
    Convert a Stock_Out_Voucher to a Sale_order with corresponding Sale_item entries.
    This will decrement product stock according to sale quantity.
    """
    try:
        voucher = models.Stock_Out_Voucher.objects.get(id=id)
    except models.Stock_Out_Voucher.DoesNotExist:
        messages.error(request, _('Voucher not found.'))
        return redirect('/stock_out_voucher')

    # Prevent duplicate conversion (session-based guard) 
    # converted_ids = request.session.get('converted_stock_out_vouchers', [])
    # if voucher.id in converted_ids:
    #     messages.warning(request, _('This voucher was already converted to a sale invoice.'))
    #     return redirect(f'/view_stock_out_invoice/{voucher.id}')

    # Require voucher has items
    items_qs = voucher.Stock_Out_Voucher_items.select_related('product').all()
    if not items_qs.exists():
        messages.error(request, _('Cannot convert an empty voucher. Please add items first.'))
        return redirect(f'/view_stock_out_invoice/{voucher.id}')

    # Validate quantities and available stock before conversion
    validation_errors = []
    for it in items_qs:
        # Quantity must be positive
        try:
            qty = int(it.quantity or 0)
        except Exception:
            qty = 0
        if qty <= 0:
            validation_errors.append(_(\
                'Item "%(prod)s" has non-positive quantity.') % {'prod': (it.product.product_name if it.product_id else '')}
            )
            continue

        # Check available stock
        prod = it.product
        try:
            available_qty = int(prod.quantity)
        except Exception:
            available_qty = prod.quantity or 0
        if available_qty < 0 or available_qty < qty:
            validation_errors.append(_(
                'Not enough stock for "%(prod)s". Available: %(avail)d, required: %(req)d'
            ) % {
                'prod': (prod.product_name if prod else ''),
                'avail': available_qty,
                'req': qty,
            })

    if validation_errors:
        for err in validation_errors:
            messages.error(request, err)
        return redirect(f'/view_stock_out_invoice/{voucher.id}')

    sale_order = models.create_sale_order(voucher.employee_id, customer_id= voucher.customer_id)
    sale_order.currency = voucher.currency
    sale_order.exchange_rate_to_base = voucher.exchange_rate_to_base or Decimal('1.0')
    sale_order.invoice_pay_method = voucher.invoice_pay_method
    sale_order.save()

    total_sum = Decimal('0.0')
    for item in items_qs:
        models.add_item_to_invoice(
            product_id=item.product_id,
            quantity=item.quantity,
            sale_price=item.unit_price,
            total_price=item.total_price,
            currency=item.currency,
            exchange_rate_to_base=item.exchange_rate_to_base,
            total_price_base=item.total_price_base,
        )
        # decrement product stock
        try:
            models.add_product_to_sale(item.product_id, item.quantity)
        except Exception:
            pass
        total_sum += Decimal(str(item.total_price))

    sale_order.grand_total = total_sum
    sale_order.total_amount = total_sum
    sale_order.total_amount_base = total_sum * Decimal(str(sale_order.exchange_rate_to_base or 1))
    sale_order.save()

    # Mark this voucher as converted in the current session
    try:
        converted_ids.append(voucher.id)
        request.session['converted_stock_out_vouchers'] = converted_ids
        request.session.modified = True
    except Exception:
        # Non-fatal: session write failure should not block successful conversion
        pass

    messages.success(request, _('Converted to sale invoice successfully.'))
    return redirect(f'/view_sale_order/{sale_order.id}')